from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
import base64
import io
import urllib.request
import urllib.error

from flask import Flask, request, render_template, redirect, url_for, session, send_file, after_this_request, jsonify
import json
import logging
import os
import re
import shutil
import secrets
import hashlib
import tempfile
import threading
import time
from urllib.parse import urlparse
from typing import List, Union, Optional
from uuid import uuid4

import numpy as np
import xgboost as xgb

# Keep app import lightweight during unit tests.
UNIT_TESTING = os.getenv("UNIT_TESTING") == "1"

import pandas as pd


try:
    import redis  # type: ignore
except Exception:  # pragma: no cover
    redis = None

try:
    from cryptography.fernet import Fernet, InvalidToken  # type: ignore
except Exception:  # pragma: no cover
    Fernet = None
    InvalidToken = Exception

try:
    # Pure-python PDF encryption.
    from pypdf import PdfReader, PdfWriter  # type: ignore
except Exception:  # pragma: no cover
    PdfReader = None
    PdfWriter = None
    logging.getLogger(__name__).warning(
        "pypdf is not installed; PDF password protection is disabled. "
        "Install it with 'pip install pypdf' (in the same venv as the app)."
    )

try:
    # Optional: AES-encrypted ZIP for sending PDF+XLSX together.
    import pyzipper  # type: ignore
except Exception:  # pragma: no cover
    pyzipper = None

try:
    # Optional: used to apply worksheet/workbook protection to XLSX.
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.workbook.protection import WorkbookProtection, hash_password  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None
    WorkbookProtection = None
    hash_password = None

try:
    # Optional but recommended for robust email validation.
    # If not installed, we fall back to a simple regex.
    from email_validator import validate_email as _ev_validate_email, EmailNotValidError
except ImportError:  # pragma: no cover
    _ev_validate_email = None

    class EmailNotValidError(Exception):
        pass

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    def load_dotenv(*_args, **_kwargs):
        return False

    logging.getLogger(__name__).warning(
        "python-dotenv is not installed; .env will not be loaded automatically. "
        "Install it with 'pip install python-dotenv' or set env vars in the OS/hosting platform."
    )

# Load secrets/config from .env (do not commit .env)
# Must run before reading env-driven globals (e.g., RUNTIME_DIR).
# Load from the project directory (same folder as this file) so it works even when
# the app is started from a different current working directory.
_PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
# IMPORTANT: never let a local .env override hosting provider environment variables.
# Render/Heroku/etc. set env vars externally; those should win.
load_dotenv(dotenv_path=os.path.join(_PROJECT_DIR, ".env"), override=False)

# ---------- RUNTIME DIR (set early for Matplotlib) ----------
RUNTIME_DIR = os.getenv("RUNTIME_DIR", tempfile.gettempdir())

if not UNIT_TESTING:
    MPLCONFIG_DIR = os.path.join(RUNTIME_DIR, "matplotlib")
    os.makedirs(MPLCONFIG_DIR, exist_ok=True)
    os.environ.setdefault("MPLCONFIGDIR", MPLCONFIG_DIR)
    os.environ.setdefault("MPLBACKEND", "Agg")

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
else:
    plt = None
    mdates = None
    mean_absolute_error = None
    mean_squared_error = None
    r2_score = None
    A4 = None
    canvas = None
    pdfmetrics = None
    TTFont = None
from werkzeug.utils import secure_filename

if not UNIT_TESTING:
    from supabase_client import get_supabase, get_supabase_admin
else:
    def get_supabase(access_token: str | None = None):
        raise RuntimeError("Supabase client is unavailable in UNIT_TESTING mode")

    def get_supabase_admin():
        raise RuntimeError("Supabase admin client is unavailable in UNIT_TESTING mode")

logging.basicConfig(level=logging.INFO)

# ---------- RUNTIME FOLDERS (writable on Render) ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(RUNTIME_DIR, "retail_forecast_uploads")
GRAPH_DIR = os.path.join(RUNTIME_DIR, "retail_forecast_graphs")
DOWNLOAD_DIR = os.path.join(RUNTIME_DIR, "retail_forecast_downloads")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(GRAPH_DIR, exist_ok=True)
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

ALLOWED_UPLOAD_EXTS = {".csv", ".xlsx", ".xls"}
DEFAULT_TREND_EPS = 0.01
DEFAULT_CLEANUP_MAX_AGE_HOURS = 24.0
MAX_FORECAST_MONTHS = 12

# ---------- PDF PASSWORD PROTECTION + ABUSE PREVENTION ----------
# These features are deliberately implemented as best-effort guards:
# - If Redis / crypto dependencies are missing, the app must not crash.
# - Password-protected sharing is rejected when secure storage is unavailable.

PDF_PASSWORD_MIN_LEN = int(os.getenv("PDF_PASSWORD_MIN_LEN", "8"))
PDF_PASSWORD_MAX_LEN = int(os.getenv("PDF_PASSWORD_MAX_LEN", "128"))
PDF_PASSWORD_TTL_SECONDS = int(os.getenv("PDF_PASSWORD_TTL_SECONDS", str(60 * 60)))  # default: 1 hour

BLOCK_DISPOSABLE_EMAILS = os.getenv("BLOCK_DISPOSABLE_EMAILS", "0").strip() in ("1", "true", "True", "yes", "YES")
DISPOSABLE_DOMAIN_BLOCKLIST = {
    d.strip().lower()
    for d in (os.getenv("DISPOSABLE_DOMAIN_BLOCKLIST") or "").split(",")
    if d.strip()
}
# Small default list to support the "optionally block" requirement without extra files.
_DEFAULT_DISPOSABLE_DOMAINS = {
    "mailinator.com",
    "10minutemail.com",
    "guerrillamail.com",
    "tempmail.com",
    "yopmail.com",
}

PWNED_PASSWORDS_ENFORCE = os.getenv("PWNED_PASSWORDS_ENFORCE", "0").strip() in ("1", "true", "True")
PWNED_PASSWORDS_TIMEOUT_SECONDS = float(os.getenv("PWNED_PASSWORDS_TIMEOUT_SECONDS", "4"))

REDIS_URL = (os.getenv("REDIS_URL") or "").strip()
PDF_PASSWORD_FERNET_KEY = (os.getenv("PDF_PASSWORD_FERNET_KEY") or "").strip()

VERIFY_EMAIL_RATE_LIMIT = int(os.getenv("VERIFY_EMAIL_RATE_LIMIT", "5"))  # per window
VERIFY_EMAIL_RATE_WINDOW_SECONDS = int(os.getenv("VERIFY_EMAIL_RATE_WINDOW_SECONDS", str(60 * 60)))

REPORT_SEND_RATE_LIMIT = int(os.getenv("REPORT_SEND_RATE_LIMIT", "10"))
REPORT_SEND_RATE_WINDOW_SECONDS = int(os.getenv("REPORT_SEND_RATE_WINDOW_SECONDS", str(15 * 60)))

RECIPIENT_COOLDOWN_SECONDS = int(os.getenv("RECIPIENT_COOLDOWN_SECONDS", str(5 * 60)))

# ---------- SETTINGS: PASSWORD CHANGE OTP + AVATAR UPLOAD ----------
PW_CHANGE_OTP_TTL_SECONDS = int(os.getenv("PW_CHANGE_OTP_TTL_SECONDS", str(10 * 60)))
PW_CHANGE_OTP_COOLDOWN_SECONDS = int(os.getenv("PW_CHANGE_OTP_COOLDOWN_SECONDS", str(60)))
PW_CHANGE_OTP_MAX_ATTEMPTS = int(os.getenv("PW_CHANGE_OTP_MAX_ATTEMPTS", str(5)))
PW_CHANGE_MIN_PASSWORD_LEN = int(os.getenv("PW_CHANGE_MIN_PASSWORD_LEN", str(8)))

COMMON_PASSWORDS = {
    "password",
    "password123",
    "admin",
    "qwerty",
    "letmein",
    "123456",
    "12345678",
    "123456789",
    "iloveyou",
    "welcome",
}


def _client_ip() -> str:
    # Prefer proxy-aware header if present (Render sets X-Forwarded-For).
    xff = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
    return xff or (request.remote_addr or "").strip() or "unknown"


def _is_disposable_email_domain(email: str) -> bool:
    if not email or "@" not in email:
        return False
    domain = email.rsplit("@", 1)[-1].strip().lower()
    if not domain:
        return False
    block = set(DISPOSABLE_DOMAIN_BLOCKLIST) | set(_DEFAULT_DISPOSABLE_DOMAINS)
    return domain in block


class _InMemoryTtlStore:
    """Tiny TTL store used when Redis isn't configured.

    This is a best-effort fallback so we don't crash in environments without Redis.
    It is per-process (not shared across instances).
    """

    def __init__(self) -> None:
        self._data: dict[str, tuple[bytes, float]] = {}

    def set(self, key: str, value: bytes, ttl_seconds: int) -> None:
        self._data[key] = (value, time.time() + float(ttl_seconds))

    def get(self, key: str) -> bytes | None:
        item = self._data.get(key)
        if not item:
            return None
        value, exp = item
        if time.time() >= exp:
            self._data.pop(key, None)
            return None
        return value

    def delete(self, key: str) -> None:
        self._data.pop(key, None)


_MEM_STORE = _InMemoryTtlStore()


def _redis_client_or_none():
    if not REDIS_URL or redis is None:
        return None
    try:
        return redis.Redis.from_url(REDIS_URL, decode_responses=False)
    except Exception:
        return None


def _fernet_or_none():
    if not PDF_PASSWORD_FERNET_KEY or Fernet is None:
        return None
    try:
        return Fernet(PDF_PASSWORD_FERNET_KEY.encode("utf-8"))
    except Exception:
        return None


def _pw_store_backend():
    r = _redis_client_or_none()
    return r if r is not None else _MEM_STORE


def _share_pw_key(owner_id: str, report_id: str, recipient_email: str) -> str:
    # Deterministic key, but opaque.
    material = f"{owner_id}|{report_id}|{recipient_email.lower().strip()}"
    return "sharepw:" + hashlib.sha256(material.encode("utf-8")).hexdigest()


def _share_pw_required_key(owner_id: str, report_id: str, recipient_email: str) -> str:
    material = f"required|{owner_id}|{report_id}|{recipient_email.lower().strip()}"
    return "sharepwreq:" + hashlib.sha256(material.encode("utf-8")).hexdigest()


def _store_share_pdf_password(owner_id: str, report_id: str, recipient_email: str, password: str) -> None:
    """Store a PDF password in short-lived encrypted storage.

    Security reasoning:
    - We never persist the password in Supabase/Postgres.
    - We store only an encrypted blob (Fernet) in Redis (or in-memory fallback).
    - We also store a separate "required" marker (no secrets) so we can prevent
      accidental unprotected delivery if the secret expires.
    """
    f = _fernet_or_none()
    if f is None:
        raise RuntimeError("password_protection_not_configured")

    # Requirement: store passwords only in short-lived encrypted storage (Redis with expiry).
    # In UNIT_TESTING we allow an in-memory fallback so tests can run without external services.
    if not UNIT_TESTING and not _redis_client_or_none():
        raise RuntimeError("password_storage_unavailable")

    backend = _pw_store_backend()
    secret_key = _share_pw_key(owner_id, report_id, recipient_email)
    required_key = _share_pw_required_key(owner_id, report_id, recipient_email)
    token = f.encrypt(password.encode("utf-8"))

    if hasattr(backend, "setex"):
        backend.setex(secret_key, PDF_PASSWORD_TTL_SECONDS, token)
        backend.setex(required_key, int(timedelta(hours=24).total_seconds()), b"1")
    else:
        backend.set(secret_key, token, PDF_PASSWORD_TTL_SECONDS)
        backend.set(required_key, b"1", int(timedelta(hours=24).total_seconds()))


def _get_share_pdf_password(owner_id: str, report_id: str, recipient_email: str) -> str | None:
    f = _fernet_or_none()
    if f is None:
        return None
    backend = _pw_store_backend()
    secret_key = _share_pw_key(owner_id, report_id, recipient_email)
    try:
        raw = backend.get(secret_key) if not hasattr(backend, "get") else backend.get(secret_key)
    except Exception:
        raw = None
    if not raw:
        return None
    try:
        return f.decrypt(raw).decode("utf-8")
    except Exception:
        return None


def _share_password_was_requested(owner_id: str, report_id: str, recipient_email: str) -> bool:
    backend = _pw_store_backend()
    required_key = _share_pw_required_key(owner_id, report_id, recipient_email)
    try:
        raw = backend.get(required_key) if not hasattr(backend, "get") else backend.get(required_key)
    except Exception:
        raw = None
    return bool(raw)


def _rate_limit_hit(key: str, limit: int, window_seconds: int) -> bool:
    """Return True if the request should be blocked (limit exceeded)."""
    backend = _pw_store_backend()
    k = "rl:" + key

    # Redis path
    if hasattr(backend, "incr") and hasattr(backend, "expire"):
        try:
            n = int(backend.incr(k))
            if n == 1:
                backend.expire(k, int(window_seconds))
            return n > int(limit)
        except Exception:
            # Fail-open to preserve availability.
            return False

    # In-memory path
    now = time.time()
    raw = backend.get(k)
    if raw is None:
        backend.set(k, b"1", int(window_seconds))
        return False
    try:
        n = int(raw.decode("utf-8"))
    except Exception:
        n = 1
    n += 1
    backend.set(k, str(n).encode("utf-8"), int(window_seconds))
    return n > int(limit)


def _cooldown_block(key: str, cooldown_seconds: int) -> bool:
    """Return True if still in cooldown; otherwise start cooldown and return False."""
    backend = _pw_store_backend()
    k = "cd:" + key
    if hasattr(backend, "set") and hasattr(backend, "setnx"):
        # Redis set-if-not-exists
        try:
            ok = backend.set(k, b"1", nx=True, ex=int(cooldown_seconds))
            return not bool(ok)
        except Exception:
            return False

    if backend.get(k):
        return True
    backend.set(k, b"1", int(cooldown_seconds))
    return False


def _cooldown_is_active(key: str) -> bool:
    """Return True if cooldown is active without mutating state."""
    backend = _pw_store_backend()
    k = "cd:" + key
    try:
        return bool(backend.get(k))
    except Exception:
        return False


def _cooldown_start(key: str, cooldown_seconds: int) -> None:
    """Start cooldown best-effort."""
    backend = _pw_store_backend()
    k = "cd:" + key
    try:
        if hasattr(backend, "set") and hasattr(backend, "setnx"):
            backend.set(k, b"1", nx=True, ex=int(cooldown_seconds))
            return
    except Exception:
        return
    try:
        backend.set(k, b"1", int(cooldown_seconds))
    except Exception:
        return


def _is_password_compromised_pwned(password: str) -> bool | None:
    """Check HaveIBeenPwned Passwords API (k-anonymity).

    Returns:
    - True: seen in breaches
    - False: not seen
    - None: check failed (network/timeout)
    """
    if not password:
        return False
    sha1 = hashlib.sha1(password.encode("utf-8")).hexdigest().upper()
    prefix, suffix = sha1[:5], sha1[5:]
    url = f"https://api.pwnedpasswords.com/range/{prefix}"
    try:
        req = urllib.request.Request(url, method="GET", headers={"User-Agent": "RetailDemandForecasting"})
        with urllib.request.urlopen(req, timeout=PWNED_PASSWORDS_TIMEOUT_SECONDS) as resp:
            body = resp.read().decode("utf-8", errors="ignore")
        for line in body.splitlines():
            if ":" not in line:
                continue
            sfx, _count = line.split(":", 1)
            if sfx.strip().upper() == suffix:
                return True
        return False
    except Exception:
        return None


def _validate_pdf_password_or_error(password: str) -> str | None:
    """Return error message if password is weak/blocked, else None.

    Security reasoning:
    - We enforce strong requirements to reduce brute-force risk.
    - We block common passwords and optionally block breached passwords.
    - We never log the password.
    """
    if password is None:
        return None
    pw = (password or "").strip()
    if not pw:
        return "PDF password cannot be empty."
    # Hard floor at 8 characters.
    min_len = max(8, int(PDF_PASSWORD_MIN_LEN))
    if len(pw) < min_len:
        return f"PDF password must be at least {min_len} characters."
    if len(pw) > PDF_PASSWORD_MAX_LEN:
        return "PDF password is too long."
    if pw.lower() in COMMON_PASSWORDS:
        return "PDF password is too common. Choose a stronger password."

    if not re.search(r"[a-z]", pw):
        return "PDF password must include a lowercase letter."
    if not re.search(r"[A-Z]", pw):
        return "PDF password must include an uppercase letter."
    if not re.search(r"\d", pw):
        return "PDF password must include a number."
    if not re.search(r"[^A-Za-z0-9]", pw):
        return "PDF password must include a symbol."

    pwned = _is_password_compromised_pwned(pw)
    if pwned is True:
        return "PDF password appears in known breaches. Choose a different password."
    if pwned is None and PWNED_PASSWORDS_ENFORCE:
        return "Unable to verify password safety right now. Please try again."

    return None


def _encrypt_pdf_bytes_if_requested(pdf_bytes: bytes, password: str | None) -> bytes:
    if not password:
        return pdf_bytes
    if PdfReader is None or PdfWriter is None:
        raise RuntimeError("pdf_encryption_unavailable")

    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)

    # Be compatible across pypdf versions.
    # Goal: ensure the document cannot be opened without the provided password.
    try:
        writer.encrypt(user_password=password)
    except TypeError:
        try:
            # Older API.
            writer.encrypt(password, use_128bit=True)
        except TypeError:
            writer.encrypt(password)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _encrypt_zip_bytes(attachments: list[tuple[str, bytes]], password: str, mode: str = "aes") -> bytes:
    """Create a password-protected ZIP containing the given attachments.

    attachments: list of (filename, file_bytes)

    mode:
    - "aes": strong encryption (WZ_AES). NOTE: Windows File Explorer can't extract AES ZIPs.
    - "zipcrypto": legacy ZipCrypto encryption. Compatible with Windows extraction.
    """
    if not password:
        raise RuntimeError("missing_zip_password")
    if pyzipper is None:
        raise RuntimeError("zip_encryption_unavailable")

    mode_norm = (mode or "").strip().lower()
    if mode_norm not in {"aes", "zipcrypto"}:
        raise RuntimeError("invalid_zip_encryption_mode")

    def _resolve_pyzipper_encryption(mode_name: str):
        # pyzipper API differs by version. Some builds expose ZIP_CRYPTO on
        # `pyzipper`, others only on `pyzipper.zipfile`.
        if mode_name == "aes":
            return getattr(pyzipper, "WZ_AES", None)

        # ZipCrypto
        if hasattr(pyzipper, "ZIP_CRYPTO"):
            return getattr(pyzipper, "ZIP_CRYPTO")
        try:
            import pyzipper.zipfile as _pzf  # type: ignore

            if hasattr(_pzf, "ZIP_CRYPTO"):
                return getattr(_pzf, "ZIP_CRYPTO")
        except Exception:
            pass

        # If ZipCrypto isn't available in this pyzipper build, return None.
        return None

    encryption = _resolve_pyzipper_encryption(mode_norm)
    if encryption is None:
        if mode_norm == "zipcrypto":
            # Prefer working delivery over hard-failing on servers with older pyzipper.
            # NOTE: Windows Explorer cannot extract AES-encrypted ZIPs; users may need 7-Zip.
            app.logger.warning("ZipCrypto encryption not available in pyzipper; falling back to AES ZIP")
            encryption = _resolve_pyzipper_encryption("aes")
        if encryption is None:
            raise RuntimeError("zip_encryption_unavailable")

    out = io.BytesIO()
    with pyzipper.AESZipFile(out, "w", compression=pyzipper.ZIP_DEFLATED, encryption=encryption) as zf:
        zf.setpassword(password.encode("utf-8"))
        for name, data in attachments:
            safe_name = os.path.basename(str(name or "").strip()) or "file"
            zf.writestr(safe_name, data)
    return out.getvalue()


def _encrypt_zip_bytes_aes(attachments: list[tuple[str, bytes]], password: str) -> bytes:
    """Backward-compatible helper: strong AES-encrypted ZIP."""
    return _encrypt_zip_bytes(attachments, password, mode="aes")


def _protect_xlsx_bytes_if_requested(xlsx_bytes: bytes, password: str | None) -> bytes:
    """Apply password protection to XLSX bytes if requested.

    Important: OpenPyXL cannot apply "file open" encryption for XLSX.
    This implements workbook/worksheet protection so edits require a password.
    """
    if not password:
        return xlsx_bytes
    if load_workbook is None or WorkbookProtection is None or hash_password is None:
        raise RuntimeError("xlsx_protection_unavailable")

    wb = load_workbook(io.BytesIO(xlsx_bytes))

    # Best-effort lock workbook structure.
    try:
        wb.security = WorkbookProtection(lockStructure=True, workbookPassword=hash_password(password))
    except Exception:
        pass

    for ws in wb.worksheets:
        try:
            ws.protection.sheet = True
            ws.protection.set_password(password)
        except Exception:
            pass

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _encrypt_xlsx_bytes_password_to_open_or_none(xlsx_bytes: bytes, password: str | None) -> bytes | None:
    """Return XLSX bytes encrypted with a password-to-open, if possible.

    This requires Windows + Microsoft Excel installed (COM automation).
    If unavailable, returns None so callers can fall back.
    """
    if not password:
        return None
    if os.name != "nt":
        return None

    try:
        import win32com.client  # type: ignore
    except Exception:
        return None

    # Write input to a temp file, then use Excel to SaveAs with password.
    with tempfile.TemporaryDirectory(prefix="xlsxpw_") as td:
        in_path = os.path.join(td, "in.xlsx")
        out_path = os.path.join(td, "out.xlsx")
        with open(in_path, "wb") as f:
            f.write(xlsx_bytes)

        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(in_path, ReadOnly=False)

            # SaveAs supports Password parameter for password-to-open.
            # FileFormat 51 is xlOpenXMLWorkbook (.xlsx)
            wb.SaveAs(out_path, FileFormat=51, Password=password)
        except Exception:
            return None
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass

        try:
            with open(out_path, "rb") as f:
                return f.read()
        except Exception:
            return None


class CacheWriteError(RuntimeError):
    """Raised when writing a cache/temp file fails."""


class InvalidResetTokens(ValueError):
    """Raised when required password reset tokens are missing or invalid."""


class TwilioSendError(RuntimeError):
    """Raised when WhatsApp sending via Twilio fails."""


@dataclass(frozen=True)
class CurrentUser:
    id: str
    email: str
    email_verified: bool


def _token_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _new_verify_token() -> str:
    # URL-safe token suitable for query string.
    return secrets.token_urlsafe(32)


def _sender_current_user_from_session() -> CurrentUser | None:
    if not session.get("logged_in"):
        return None
    user_id = (session.get("user_id") or "").strip()
    email = (session.get("user_email") or "").strip().lower()
    email_verified = bool(session.get("email_verified"))
    if not user_id or not email:
        return None
    return CurrentUser(id=user_id, email=email, email_verified=email_verified)


def _compute_email_verified_from_user_obj(user_obj) -> bool:
    return bool(_resp_get(user_obj, "email_confirmed_at") or _resp_get(user_obj, "confirmed_at"))


def _refresh_sender_verified_state_best_effort() -> None:
    """Refresh session['email_verified'] based on Supabase user object (best-effort).

    This keeps policy enforcement stable even if the login response didn't include
    the confirmed fields in a consistent shape.
    """
    if UNIT_TESTING:
        return
    if not session.get("logged_in"):
        return
    access_token = (session.get("access_token") or "").strip()
    if not access_token:
        return

    try:
        sb_user = get_supabase(access_token)
        # supabase-py has had different signatures; support both.
        try:
            user_resp = sb_user.auth.get_user(access_token)
        except TypeError:
            user_resp = sb_user.auth.get_user()
        user_obj = _resp_get(user_resp, "user") or _resp_get(user_resp, "data") or user_resp
        if user_obj is None:
            return
        session["email_verified"] = _compute_email_verified_from_user_obj(user_obj)
    except Exception:
        # Do not block user actions due to refresh failures.
        return


def _supabase_table_get_single(sb, table: str, filters: dict) -> dict | None:
    q = sb.table(table).select("*")
    for k, v in filters.items():
        q = q.eq(k, v)
    res = q.limit(1).execute()
    data = getattr(res, "data", None)
    if isinstance(data, list) and data:
        return data[0]
    if isinstance(data, dict):
        return data
    return None


def _trusted_email_is_verified(owner_id: str, email: str, access_token: str) -> bool:
    if UNIT_TESTING:
        return False
    recipient = (email or "").strip().lower()
    if not recipient:
        return False

    def _query(sb) -> dict | None:
        # Prefer exact match, then case-insensitive match (if supported by client).
        row = _supabase_table_get_single(sb, "trusted_emails", {"owner_id": owner_id, "email": recipient})
        if row:
            return row
        try:
            res = (
                sb.table("trusted_emails")
                .select("*")
                .eq("owner_id", owner_id)
                .ilike("email", recipient)
                .limit(1)
                .execute()
            )
            data = getattr(res, "data", None)
            if isinstance(data, list) and data:
                return data[0]
            if isinstance(data, dict):
                return data
        except Exception:
            return None
        return None

    # First try user-scoped client.
    try:
        sb = get_supabase(access_token)
        row = _query(sb)
        return bool(row and row.get("is_verified"))
    except Exception:
        # If the access token is expired/invalid on Render, fall back to the admin client.
        # We still filter by owner_id, so this remains scoped.
        try:
            sb_admin = get_supabase_admin()
            row = _query(sb_admin)
            return bool(row and row.get("is_verified"))
        except Exception:
            return False


def _report_share_is_verified(owner_id: str, report_id: str, recipient_email: str, access_token: str) -> bool:
    if UNIT_TESTING:
        return False
    recipient = (recipient_email or "").strip().lower()
    if not recipient:
        return False

    def _query(sb) -> dict | None:
        row = _supabase_table_get_single(
            sb,
            "report_shares",
            {"owner_id": owner_id, "report_id": report_id, "recipient_email": recipient},
        )
        if row:
            return row
        # Case-insensitive fallback if supported.
        try:
            res = (
                sb.table("report_shares")
                .select("*")
                .eq("owner_id", owner_id)
                .eq("report_id", report_id)
                .ilike("recipient_email", recipient)
                .limit(1)
                .execute()
            )
            data = getattr(res, "data", None)
            if isinstance(data, list) and data:
                return data[0]
            if isinstance(data, dict):
                return data
        except Exception:
            return None
        return None

    try:
        sb = get_supabase(access_token)
        row = _query(sb)
    except Exception:
        try:
            sb_admin = get_supabase_admin()
            row = _query(sb_admin)
        except Exception:
            row = None

    if not row:
        return False
    if not row.get("is_verified"):
        return False
    # NOTE:
    # `expires_at` is used to expire the *verification token*, not the verification itself.
    # Once a share is verified, we must not block delivery just because the original
    # token expired.
    return True


def _report_share_recipient_is_verified_any(owner_id: str, recipient_email: str, access_token: str) -> bool:
    """Return True if recipient_email has ANY verified share row for this owner.

    This prevents surprising behavior where a recipient is forced to re-verify for
    every newly generated report_id.
    """
    if UNIT_TESTING:
        return False
    recipient = (recipient_email or "").strip().lower()
    if not recipient:
        return False

    def _query_any(sb) -> bool:
        res = (
            sb.table("report_shares")
            .select("id,is_verified")
            .eq("owner_id", owner_id)
            .eq("is_verified", True)
            .limit(1)
        )
        # Prefer exact; fall back to ilike when available.
        try:
            res = res.eq("recipient_email", recipient)
        except Exception:
            pass
        try:
            res2 = res.execute()
            rows = getattr(res2, "data", None) or []
            if rows:
                return True
        except Exception:
            pass

        try:
            res3 = (
                sb.table("report_shares")
                .select("id,is_verified")
                .eq("owner_id", owner_id)
                .eq("is_verified", True)
                .ilike("recipient_email", recipient)
                .limit(1)
                .execute()
            )
            rows = getattr(res3, "data", None) or []
            return bool(rows)
        except Exception:
            return False

    # First user-scoped, then admin fallback.
    try:
        sb = get_supabase(access_token)
        if _query_any(sb):
            return True
    except Exception:
        pass

    try:
        sb_admin = get_supabase_admin()
        return _query_any(sb_admin)
    except Exception:
        return False


def can_send_report(user: CurrentUser, recipient_email: str, report_id: str | None) -> bool:
    """Authorization policy for delivering a report to recipient_email.

    Rules (must satisfy all of these):
    - Sender authenticated AND email_verified
    - Recipient email syntax is valid (done before calling this)
    - Allow if ANY ONE is true:
        1) recipient == sender verified email
        2) recipient in sender.trusted_emails AND verified
        3) recipient verified via one-time share token for this report
    """
    if not user or not user.email_verified:
        return False

    recipient = (recipient_email or "").strip().lower()
    if not recipient:
        return False

    if recipient == user.email:
        return True

    access_token = (session.get("access_token") or "").strip()
    if not access_token:
        return False

    if _trusted_email_is_verified(user.id, recipient, access_token):
        return True

    if report_id and _report_share_is_verified(user.id, report_id, recipient, access_token):
        return True

    # If a recipient verified a previous share for this owner, allow delivery for
    # new report_ids without forcing repeated verification.
    if _report_share_recipient_is_verified_any(user.id, recipient, access_token):
        return True

    return False

@dataclass(frozen=True)
class AppConfig:
    """Application configuration parsed once at startup from environment variables."""

    app_base_url: str
    email_redirect_base_url: str
    flask_secret_key: str
    trend_eps: float
    cleanup_max_age_hours: float
    supabase_url: str
    supabase_anon_key: str
    public_https_base_url: str

    @staticmethod
    def from_env() -> "AppConfig":
        def _first_non_empty_str(*values: str | None) -> str | None:
            for v in values:
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    return s
            return None

        def _safe_float_env(name: str, default: float) -> float:
            raw = os.getenv(name)
            if raw is None:
                return float(default)
            try:
                value = float(str(raw).strip())
                if not np.isfinite(value):
                    return float(default)
                return value
            except Exception:
                return float(default)

        # Prefer explicit APP_BASE_URL, but fall back to Render's public service URL
        # so production deploys don't accidentally embed localhost links in emails.
        app_base_url = (
            _first_non_empty_str(
                os.getenv("APP_BASE_URL"),
                os.getenv("RENDER_EXTERNAL_URL"),
                "http://127.0.0.1:5000",
            )
            or "http://127.0.0.1:5000"
        ).rstrip("/")
        # Base URL to embed in emailed links (confirmation + password reset).
        # Set this to your deployed domain in production.
        email_redirect_base_url = (
            _first_non_empty_str(
                os.getenv("EMAIL_REDIRECT_BASE_URL"),
                os.getenv("PUBLIC_BASE_URL"),
                app_base_url,
            )
            or app_base_url
        ).rstrip("/")
        flask_secret_key = os.getenv("FLASK_SECRET_KEY", "retail_secret_key")
        trend_eps = _safe_float_env("TREND_EPS", DEFAULT_TREND_EPS)
        cleanup_max_age_hours = _safe_float_env("CLEANUP_MAX_AGE_HOURS", DEFAULT_CLEANUP_MAX_AGE_HOURS)

        # Support both server-side and Vite-style variable names.
        supabase_url = (os.getenv("SUPABASE_URL") or os.getenv("VITE_SUPABASE_URL") or "").strip()
        supabase_anon_key = (os.getenv("SUPABASE_ANON_KEY") or os.getenv("VITE_SUPABASE_ANON_KEY") or "").strip()

        # Public HTTPS base URL for Twilio MediaUrl fetching.
        # Examples: https://<your-ngrok>.ngrok-free.app
        public_https_base_url = (
            os.getenv("PUBLIC_HTTPS_BASE_URL")
            or os.getenv("PUBLIC_BASE_URL")
            or os.getenv("EMAIL_REDIRECT_BASE_URL")
            or os.getenv("APP_BASE_URL")
            or ""
        ).strip().rstrip("/")

        return AppConfig(
            app_base_url=app_base_url,
            email_redirect_base_url=email_redirect_base_url,
            flask_secret_key=flask_secret_key,
            trend_eps=float(trend_eps),
            cleanup_max_age_hours=float(cleanup_max_age_hours),
            supabase_url=supabase_url,
            supabase_anon_key=supabase_anon_key,
            public_https_base_url=public_https_base_url,
        )


CONFIG = AppConfig.from_env()


app = Flask(__name__)
app.static_folder = GRAPH_DIR
app.secret_key = CONFIG.flask_secret_key

# Session cookie defaults (helps login sessions behave consistently)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)

# Base URL for auth redirects
APP_BASE_URL = CONFIG.app_base_url


def _get_env_float(name: str, default: float) -> float:
    """Parse a float environment variable safely.

    - Never raises due to invalid user-provided env var values.
    - Avoids parsing at import time by being called from request-time helpers.
    """
    raw = os.getenv(name)
    if raw is None:
        return float(default)
    try:
        value = float(str(raw).strip())
        if not np.isfinite(value):
            return float(default)
        return value
    except Exception:
        return float(default)


def _get_env_int(name: str, default: int) -> int:
    """Parse an int environment variable safely.

    - Never raises due to invalid user-provided env var values.
    - Avoids parsing at import time by being called from configuration helpers.
    """
    raw = os.getenv(name)
    if raw is None:
        return int(default)
    try:
        value = int(str(raw).strip())
        return value
    except Exception:
        return int(default)


def _get_env_bool(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return bool(default)
    v = str(raw).strip().lower()
    if v in ("1", "true", "yes", "y", "on"):
        return True
    if v in ("0", "false", "no", "n", "off"):
        return False
    return bool(default)


_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def _is_valid_email_address(value: str | None) -> bool:
    if not value:
        return False
    v = str(value).strip()
    if not v or len(v) > 254:
        return False
    return _EMAIL_RE.match(v) is not None


def _normalize_and_validate_email_for_sending(value: str | None) -> tuple[str | None, str | None]:
    """Validate and normalize an email address for outbound sending.

    - Uses `email-validator` when available (better parsing/normalization).
    - Falls back to a conservative regex check if not installed.
    - Optional deliverability check (DNS) can be enabled via env var:
        EMAIL_DELIVERABILITY_CHECK=1
    """
    if value is None:
        return None, "Please enter a valid email address."
    raw = str(value).strip()
    if not raw:
        return None, "Please enter a valid email address."

    if _ev_validate_email is None:
        return (raw, None) if _is_valid_email_address(raw) else (None, "Please enter a valid email address.")

    check_deliverability = _get_env_bool("EMAIL_DELIVERABILITY_CHECK", False)
    try:
        res = _ev_validate_email(raw, check_deliverability=bool(check_deliverability))
        normalized = (getattr(res, "email", None) or raw).strip()
        if not normalized or len(normalized) > 254:
            return None, "Please enter a valid email address."
        return normalized, None
    except EmailNotValidError:
        return None, "Please enter a valid email address."


def _get_sendgrid_config() -> dict:
    """Return SendGrid API config from environment.

    Required:
    - SENDGRID_API_KEY
    - SENDGRID_FROM
    """
    api_key = (os.getenv("SENDGRID_API_KEY") or "").strip()
    from_email = (os.getenv("SENDGRID_FROM") or "").strip()
    if api_key and not api_key.startswith("SG."):
        app.logger.warning("SENDGRID_API_KEY does not look like a SendGrid API key (expected prefix 'SG.').")
    return {
        "api_key": api_key,
        "from_email": from_email,
    }


def _get_twilio_config() -> dict:
    """Return Twilio WhatsApp config from environment.

    Required:
    - TWILIO_ACCOUNT_SID
    - TWILIO_AUTH_TOKEN

    Optional:
    - TWILIO_WHATSAPP_FROM (defaults to Twilio sandbox number)
    """
    account_sid = (os.getenv("TWILIO_ACCOUNT_SID") or "").strip()
    auth_token = (os.getenv("TWILIO_AUTH_TOKEN") or "").strip()
    whatsapp_from = (os.getenv("TWILIO_WHATSAPP_FROM") or "whatsapp:+14155238886").strip()

    if account_sid and not account_sid.startswith("AC"):
        app.logger.warning("TWILIO_ACCOUNT_SID does not look like an Account SID (expected prefix 'AC').")
    if whatsapp_from and not whatsapp_from.startswith("whatsapp:+"):
        app.logger.warning("TWILIO_WHATSAPP_FROM must start with 'whatsapp:+'.")

    return {
        "account_sid": account_sid,
        "auth_token": auth_token,
        "whatsapp_from": whatsapp_from,
    }


_WHATSAPP_E164_RE = re.compile(r"^\+[1-9]\d{7,14}$")
_WHATSAPP_IN_RE = re.compile(r"^\+91\d{10}$")


def _is_valid_whatsapp_number(value: str | None) -> bool:
    """Validate WhatsApp number input.

    UX requirement calls out +91XXXXXXXXXX, but we accept generic E.164 too.
    """
    if not value:
        return False
    v = str(value).strip()
    if not v:
        return False
    return _WHATSAPP_IN_RE.match(v) is not None or _WHATSAPP_E164_RE.match(v) is not None


def _public_https_base_url_or_none() -> str | None:
    base = (CONFIG.public_https_base_url or "").strip().rstrip("/")
    if not base:
        # On deployed environments (e.g., Render + custom domains), we can often infer the
        # correct public HTTPS base URL from the incoming request.
        try:
            inferred = (getattr(request, "url_root", "") or "").strip().rstrip("/")
        except Exception:
            inferred = ""
        if inferred.lower().startswith("https://"):
            return inferred
        return None
    if not base.lower().startswith("https://"):
        return None
    return base


def _parse_report_id(report_id: str | None) -> str | None:
    if not report_id:
        return None
    v = str(report_id).strip()
    # report_id is stored from uuid4(), typically in canonical UUID form.
    # Only allow UUID-like chars to avoid path traversal.
    if not re.fullmatch(r"[0-9a-fA-F-]{32,36}", v):
        return None
    return v


def _build_excel_report_file(report: dict, path: str) -> None:
    """Generate the Excel report at the given path."""
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        pd.DataFrame([report.get('metrics', {})]).to_excel(writer, sheet_name="Metrics", index=False)
        pd.DataFrame([report.get('insights', {})]).to_excel(writer, sheet_name="Insights", index=False)
        pd.DataFrame([report.get('trends', {})]).to_excel(writer, sheet_name="Trends", index=False)
        pd.DataFrame(report.get('future_forecast_table', [])).to_excel(
            writer, sheet_name="Future_Forecast", index=False
        )
        pd.DataFrame(report.get('category_cost_table', [])).to_excel(
            writer, sheet_name="Monthwise_Category_Cost", index=False
        )


def _ensure_public_report_files(report_id: str, report: dict) -> tuple[str, str]:
    """Ensure public-facing PDF/XLSX files exist for Twilio MediaUrl fetching."""
    rid = _parse_report_id(report_id)
    if not rid:
        raise RuntimeError("invalid_report_id")

    pdf_path = os.path.join(DOWNLOAD_DIR, f"public_{rid}.pdf")
    xlsx_path = os.path.join(DOWNLOAD_DIR, f"public_{rid}.xlsx")

    # Generate if missing (best-effort idempotent).
    if not os.path.exists(pdf_path):
        _build_pdf_report_file(report, pdf_path)
    if not os.path.exists(xlsx_path):
        _build_excel_report_file(report, xlsx_path)
    return pdf_path, xlsx_path


def _twilio_error_to_user_message(err: Exception) -> str:
    """Map Twilio exceptions to safe, user-friendly messages.

    Note: Do not include secrets or full Twilio error payloads in user responses.
    """

    # Prefer structured info when Twilio SDK provides it.
    twilio_code = getattr(err, "code", None)
    twilio_status = getattr(err, "status", None)
    twilio_msg = getattr(err, "msg", None)
    msg = str(twilio_msg or err or "")
    lower = msg.lower()

    # WhatsApp sandbox join requirement.
    # Twilio commonly returns code 63016 when the recipient hasn't joined the sandbox.
    if twilio_code == 63016 or "63016" in lower or ("sandbox" in lower and "join" in lower):
        return "WhatsApp sandbox not joined. Please join the Twilio WhatsApp Sandbox first, then try again."

    # Invalid destination format.
    if twilio_code in (21211, 21614) or "not a valid phone number" in lower or "is not a valid" in lower:
        return "Invalid WhatsApp number. Use format like +91XXXXXXXXXX."

    # Authentication/authorization failures.
    # Twilio uses 20003 for invalid AccountSid/AuthToken.
    if twilio_code == 20003 or "authenticate" in lower or str(twilio_status) == "401" or "http 401" in lower:
        return "WhatsApp service authentication failed. Verify TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN on the server."

    # Media fetch failures (Twilio can't reach your public file URLs).
    if "media" in lower and ("url" in lower or "fetch" in lower or "retrieve" in lower or "https" in lower):
        return "WhatsApp send failed because Twilio could not fetch the report files. Ensure PUBLIC_HTTPS_BASE_URL is a public HTTPS domain and /files/... is reachable."

    return "Failed to send WhatsApp message. Please try again later."


def _send_whatsapp_media_twilio(to_number_e164: str, media_url: str, body: str) -> str:
    cfg = _get_twilio_config()
    if not cfg.get("account_sid") or not cfg.get("auth_token"):
        raise RuntimeError("missing_twilio_config")

    # Import lazily so UNIT_TESTING doesn't require twilio installed.
    try:
        from twilio.rest import Client  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError("missing_twilio_sdk") from e

    to = f"whatsapp:{to_number_e164.strip()}"
    from_ = cfg.get("whatsapp_from")
    client = Client(cfg["account_sid"], cfg["auth_token"])

    # Twilio fetches media from the URL; must be HTTPS and publicly reachable.
    msg = client.messages.create(
        to=to,
        from_=from_,
        body=body,
        media_url=[media_url],
    )
    sid = getattr(msg, "sid", None) or ""
    if not sid:
        raise TwilioSendError("twilio_no_sid")
    return sid


def _send_email_with_pdf_attachment_sendgrid(
    to_email: str,
    subject: str,
    body: str,
    pdf_bytes: bytes,
    filename: str,
) -> None:
    cfg = _get_sendgrid_config()
    if not cfg.get("api_key") or not cfg.get("from_email"):
        raise RuntimeError("missing_sendgrid_config")

    payload = {
        "personalizations": [{"to": [{"email": to_email}]}],
        "from": {"email": cfg["from_email"]},
        "subject": subject,
        "content": [{"type": "text/plain", "value": body}],
        "attachments": [
            {
                "content": base64.b64encode(pdf_bytes).decode("ascii"),
                "type": "application/pdf",
                "filename": filename,
                "disposition": "attachment",
            }
        ],
    }

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.sendgrid.com/v3/mail/send",
        data=data,
        method="POST",
        headers={
            "Authorization": f"Bearer {cfg['api_key']}",
            "Content-Type": "application/json",
        },
    )

    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            status = getattr(resp, "status", None)
            if status not in (200, 202):
                raise RuntimeError("sendgrid_failed")
    except urllib.error.HTTPError as e:
        # SendGrid returns a JSON body with details; log it for debugging but keep user-facing errors generic.
        try:
            body_bytes = e.read()
            body_text = (body_bytes or b"").decode("utf-8", errors="replace")
        except Exception:
            body_text = ""
        app.logger.warning(
            "SendGrid HTTPError code=%s body=%s",
            getattr(e, "code", "unknown"),
            (body_text[:2000] if body_text else ""),
        )
        raise RuntimeError("sendgrid_failed") from e
    except urllib.error.URLError as e:
        app.logger.warning("SendGrid URLError: %s", str(e))
        raise RuntimeError("sendgrid_failed") from e


def _send_email_with_attachments_sendgrid(
    to_email: str,
    subject: str,
    body: str,
    attachments: list[tuple[str, str, bytes]],
) -> None:
    """Send an email with multiple attachments via SendGrid.

    attachments: list of (filename, mime_type, content_bytes)
    """
    cfg = _get_sendgrid_config()
    if not cfg.get("api_key") or not cfg.get("from_email"):
        raise RuntimeError("missing_sendgrid_config")

    payload = {
        "personalizations": [{"to": [{"email": to_email}]}],
        "from": {"email": cfg["from_email"]},
        "subject": subject,
        "content": [{"type": "text/plain", "value": body}],
        "attachments": [
            {
                "content": base64.b64encode(content).decode("ascii"),
                "type": mime,
                "filename": filename,
                "disposition": "attachment",
            }
            for (filename, mime, content) in (attachments or [])
        ],
    }

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.sendgrid.com/v3/mail/send",
        data=data,
        method="POST",
        headers={
            "Authorization": f"Bearer {cfg['api_key']}",
            "Content-Type": "application/json",
        },
    )

    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            status = getattr(resp, "status", None)
            if status not in (200, 202):
                raise RuntimeError("sendgrid_failed")
    except urllib.error.HTTPError as e:
        try:
            body_bytes = e.read()
            body_text = (body_bytes or b"").decode("utf-8", errors="replace")
        except Exception:
            body_text = ""
        app.logger.warning(
            "SendGrid HTTPError (attachments) code=%s body=%s",
            getattr(e, "code", "unknown"),
            (body_text[:2000] if body_text else ""),
        )
        raise RuntimeError("sendgrid_failed") from e
    except urllib.error.URLError as e:
        app.logger.warning("SendGrid URLError (attachments): %s", str(e))
        raise RuntimeError("sendgrid_failed") from e


def _build_pdf_report_file(report: dict, path: str) -> None:
    """Generate the PDF report at the given path."""
    c = canvas.Canvas(path, pagesize=A4)
    width, height = A4

    # Prefer a Unicode-capable TrueType font so the Rupee symbol (₹) renders correctly.
    FONT_REGULAR, FONT_BOLD = _get_pdf_font_names()

    top_margin = 50
    bottom_margin = 50
    left = 50
    right = 50

    def new_page(font_size: int = 11) -> float:
        c.showPage()
        c.setLineWidth(1)
        c.setFont(FONT_REGULAR, font_size)
        return height - top_margin

    def ensure_space(y: float, needed: float, font_size: int = 11) -> float:
        if y - needed < bottom_margin:
            return new_page(font_size=font_size)
        return y

    def draw_section_title(y: float, title: str) -> float:
        y = ensure_space(y, 24, font_size=11)
        c.setFont(FONT_BOLD, 12)
        c.drawString(left, y, title)
        c.setFont(FONT_REGULAR, 11)
        return y - 18

    def draw_centered_title(y: float, title: str) -> float:
        y = ensure_space(y, 30, font_size=11)
        c.setFont(FONT_BOLD, 16)
        c.drawCentredString(width / 2, y, title)
        c.setFont(FONT_REGULAR, 11)
        return y - 28

    def draw_kv_lines(y: float, items: dict, line_height: float = 16) -> float:
        for k, v in items.items():
            y = ensure_space(y, line_height)
            c.drawString(left, y, f"{k}: {v}")
            y -= line_height
        return y

    def draw_table(
        y: float,
        headers: list[str],
        rows: list[list[str]],
        col_widths: list[float],
        row_h: float = 22,
        grid_line_width: float = 2.2,
        outer_line_width: float = 2.8,
    ) -> float:
        table_w = sum(col_widths)

        def _baseline(y_top: float, h: float, font_size: int) -> float:
            return y_top - h + (h - font_size) / 2 + 3

        def _draw_header(y_top: float) -> float:
            c.setFont(FONT_BOLD, 11)
            x = left
            for header, w in zip(headers, col_widths):
                c.rect(x, y_top - row_h, w, row_h, stroke=1, fill=0)
                c.drawCentredString(x + (w / 2), _baseline(y_top, row_h, 11), str(header))
                x += w
            return y_top - row_h

        def _draw_row(y_top: float, row: list[str]) -> float:
            c.setFont(FONT_REGULAR, 11)
            x = left
            for cell, w in zip(row, col_widths):
                c.rect(x, y_top - row_h, w, row_h, stroke=1, fill=0)
                c.drawCentredString(x + (w / 2), _baseline(y_top, row_h, 11), str(cell))
                x += w
            return y_top - row_h

        c.setLineWidth(grid_line_width)
        min_needed = row_h * 2
        y = ensure_space(y, min_needed)

        segment_top = y
        segment_rows = 0
        y = _draw_header(y)

        for row in rows:
            if y - row_h < bottom_margin:
                segment_h = row_h * (1 + segment_rows)
                c.setLineWidth(outer_line_width)
                c.rect(left, segment_top - segment_h, table_w, segment_h, stroke=1, fill=0)
                c.setLineWidth(grid_line_width)

                y = new_page(font_size=11)
                y = ensure_space(y, min_needed)
                segment_top = y
                segment_rows = 0
                y = _draw_header(y)

            y = _draw_row(y, row)
            segment_rows += 1

        segment_h = row_h * (1 + segment_rows)
        c.setLineWidth(outer_line_width)
        c.rect(left, segment_top - segment_h, table_w, segment_h, stroke=1, fill=0)

        c.setLineWidth(1)
        c.setFont(FONT_REGULAR, 11)
        return y - 10

    # ---------- PAGE 1: Title + Graphs first ----------
    c.setFont(FONT_BOLD, 14)
    c.drawString(left, height - 50, "Retail Demand Forecasting Report")

    y = height - 80
    image_w = width - left - right
    image_h = 240
    gap = 18

    past_image = report.get("past_image")
    future_image = report.get("future_image")
    months = int(report.get("months") or 0)

    if past_image:
        y = ensure_space(y, image_h + gap + 18)
        c.setFont(FONT_BOLD, 12)
        c.drawString(left, y, "Actual vs Predicted Sales (Weekly Averages)")
        y -= 16
        c.drawImage(
            os.path.join(GRAPH_DIR, past_image),
            left,
            y - image_h,
            image_w,
            image_h,
            preserveAspectRatio=True,
            anchor='c',
        )
        y -= image_h + gap

    if future_image:
        y = ensure_space(y, image_h + gap + 18)
        title = "Future Sales Forecast"
        if months:
            title = f"Future Sales Forecast ({months} Month{'s' if months > 1 else ''})"
        c.setFont(FONT_BOLD, 12)
        c.drawString(left, y, title)
        y -= 16
        c.drawImage(
            os.path.join(GRAPH_DIR, future_image),
            left,
            y - image_h,
            image_w,
            image_h,
            preserveAspectRatio=True,
            anchor='c',
        )
        y -= image_h + gap

    # Start details after graphs
    y = new_page(font_size=11)

    insights = report.get('insights') or {}
    if insights:
        y = draw_section_title(y, "Business Insights & Recommendation")
        y = draw_kv_lines(y, insights)
        y -= 8

    trends = report.get('trends') or {}
    if trends:
        y = draw_section_title(y, "Trend Summary")
        y = draw_kv_lines(y, trends, line_height=15)
        y -= 8

    cat_rows = report.get('category_cost_table') or []
    if cat_rows:
        y = new_page(font_size=11)
        section_title = "Category Cost & Stock Analysis (Forecast Months)"
        y = draw_centered_title(y, section_title)
        headers = ["Month", "Product Category", "Units Sold", "Total Cost (₹)"]
        rows = []
        for r in cat_rows:
            m = r.get('month') or ""
            rows.append([
                str(m),
                str(r.get('product_category', '')),
                str(r.get('total_units_sold', '')),
                f"₹ {r.get('total_cost', '')}",
            ])
        col_widths = [70, width - left - right - (70 + 90 + 120), 90, 120]
        y = draw_table(y, headers, rows, col_widths=col_widths)

    c.save()

# ---------- LOAD MODEL ----------
MODEL_PATH = os.getenv("MODEL_PATH", os.path.join(BASE_DIR, "model", "xgb_model.json"))
if os.getenv("SKIP_MODEL_LOAD") == "1":
    model = None
else:
    if not os.path.exists(MODEL_PATH):
        raise RuntimeError(
            f"Missing model file: {MODEL_PATH}. "
            "Commit model/xgb_model.json to the deployed branch, or set MODEL_PATH."
        )

    class _XgbBoosterPredictor:
        def __init__(self, booster: "xgb.Booster") -> None:
            self._booster = booster

        def predict(self, X, validate_features: bool = True):
            # Preserve existing API usage: callers pass a pandas DataFrame.
            if isinstance(X, xgb.DMatrix):
                dm = X
            else:
                feature_names = list(X.columns) if hasattr(X, "columns") else None
                dm = xgb.DMatrix(X, feature_names=feature_names)
            try:
                return self._booster.predict(dm, validate_features=validate_features)
            except TypeError:
                return self._booster.predict(dm)

    booster = xgb.Booster()
    booster.load_model(MODEL_PATH)
    model = _XgbBoosterPredictor(booster)


# XGBoost can raise "data did not contain feature names" on some deployments
# depending on version + pandas integration. Use a safe wrapper that disables
# feature validation when supported.
FEATURE_COLS = ["lag_1", "lag_7", "roll_mean_7", "roll_mean_30"]


def _resp_get(obj, key: str):
    """Safely read a key from supabase-py responses across versions."""
    if isinstance(obj, dict):
        return obj.get(key)
    return getattr(obj, key, None)


def _get_identities_count(user_obj) -> int | None:
    """Best-effort read Supabase user.identities count across versions."""
    if user_obj is None:
        return None
    try:
        identities = _resp_get(user_obj, "identities")
        if identities is None and isinstance(user_obj, dict):
            identities = user_obj.get("identities")
        if identities is None:
            return None
        if isinstance(identities, (list, tuple)):
            return len(identities)
        # Some clients might expose identities as an iterable-like
        try:
            return len(list(identities))
        except Exception:
            return None
    except Exception:
        return None


def _parse_iso_datetime(value: str | None) -> datetime | None:
    """Parse an ISO8601-ish datetime string safely (best-effort)."""
    if not value:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Python's fromisoformat doesn't accept 'Z' suffix.
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None


def _looks_like_existing_user_after_signup(user_obj) -> bool:
    """Heuristic: True if sign_up appears to have returned an existing account.

    Some Supabase Auth configurations return a success-like response for existing
    emails (to reduce account enumeration), but do not send a confirmation email.
    We detect that case to display a clear "already registered" message.

    We keep this conservative to avoid misclassifying legitimate new signups.
    """
    if user_obj is None:
        return False

    identities_count = _get_identities_count(user_obj)
    if identities_count is None or identities_count > 0:
        return False

    # If the user is already confirmed, it must be an existing account.
    if _resp_get(user_obj, "email_confirmed_at") or _resp_get(user_obj, "confirmed_at"):
        return True

    created_at = _resp_get(user_obj, "created_at")
    created_dt = _parse_iso_datetime(created_at)
    if not created_dt:
        return False

    now = datetime.now(timezone.utc)
    age_s = (now - created_dt).total_seconds()
    # If the account is older than a few minutes and identities are empty, it's
    # very likely an existing account (not a fresh sign-up).
    return age_s > 300


def _first_non_empty(*values: str | None) -> str | None:
    for v in values:
        if v is None:
            continue
        v = str(v).strip()
        if v:
            return v
    return None


_PASSWORD_MIN_LEN = 8
_PASSWORD_UPPER_RE = re.compile(r"[A-Z]")
_PASSWORD_DIGIT_RE = re.compile(r"\d")
_PASSWORD_SPECIAL_RE = re.compile(r"[^A-Za-z0-9]")


def _password_policy_payload() -> dict:
    """Password policy rendered to templates for client-side validation.

    Keep this as the single source of truth so server + client can't drift.
    """
    return {
        "min_len": _PASSWORD_MIN_LEN,
        "upper_pattern": _PASSWORD_UPPER_RE.pattern,
        "digit_pattern": _PASSWORD_DIGIT_RE.pattern,
        "special_pattern": _PASSWORD_SPECIAL_RE.pattern,
        "labels": {
            "len": "At least 8 characters",
            "upper": "One capital letter (A-Z)",
            "digit": "One number (0-9)",
            "special": "One special character (e.g., !@#$)",
        },
    }


def _sanitize_public_error_message(msg: str | None, limit: int = 300) -> str:
    """Best-effort sanitize exception text for displaying to end users."""
    if not msg:
        return ""
    s = str(msg)
    s = s.replace("\r", " ").replace("\n", " ")
    # Strip some common wrapper prefixes from supabase-py
    s = re.sub(r"AuthApiError\([^)]*\):\s*", "", s)
    s = re.sub(r"APIError\([^)]*\):\s*", "", s)
    s = re.sub(r"HTTPError\([^)]*\):\s*", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > limit:
        s = s[:limit].rstrip() + "…"
    return s


def _password_missing_requirements(password: str) -> list[str]:
    missing: list[str] = []
    if len(password) < _PASSWORD_MIN_LEN:
        missing.append("at least 8 characters")
    if _PASSWORD_UPPER_RE.search(password) is None:
        missing.append("one capital letter (A-Z)")
    if _PASSWORD_DIGIT_RE.search(password) is None:
        missing.append("one number (0-9)")
    if _PASSWORD_SPECIAL_RE.search(password) is None:
        missing.append("one special character (e.g., !@#$)")
    return missing


def _password_strength_error(password: str | None) -> str | None:
    """Return a user-facing error string if password is weak, else None.

    Rules:
    - length >= 8
    - at least 1 uppercase letter
    - at least 1 number
    - at least 1 special character (any non-alphanumeric)
    """
    pw = (password or "")
    if not pw:
        return "Please enter a password."

    missing = _password_missing_requirements(pw)
    if not missing:
        return None

    if missing == ["at least 8 characters"]:
        return "Password is weak: must be at least 8 characters."

    return "Password is weak. Missing: " + ", ".join(missing) + "."


def _friendly_register_error(err: Exception) -> str:
    """Map Supabase sign_up errors to consistent user-friendly messages."""
    raw_msg = _sanitize_public_error_message(str(err))
    msg_lower = raw_msg.lower()

    # Duplicate email (wording varies by supabase-py / GoTrue)
    if (
        "already registered" in msg_lower
        or "user already registered" in msg_lower
        or "already exists" in msg_lower
        or "email already" in msg_lower
        or "email" in msg_lower and "in use" in msg_lower
    ):
        return "This email is already registered. Reset password instead."

    # Weak password (wording varies)
    if (
        "weak_password" in msg_lower
        or ("password" in msg_lower and "weak" in msg_lower)
        or ("password" in msg_lower and "strength" in msg_lower)
        or ("password" in msg_lower and "at least" in msg_lower)
        or ("password" in msg_lower and "short" in msg_lower)
    ):
        return "Password is weak: must be at least 8 characters and include 1 capital letter (A-Z), 1 number (0-9), and 1 special character (e.g., !@#$)."

    if "invalid email" in msg_lower:
        return "Please enter a valid email address."

    # Rate limiting / abuse protection
    if (
        "rate limit" in msg_lower
        or "too many requests" in msg_lower
        or "over_email_send_rate_limit" in msg_lower
        or "429" in msg_lower
    ):
        return "Too many attempts. Please wait and try again later."

    if "missing supabase config" in msg_lower:
        return "Registration is temporarily unavailable due to a server configuration issue."

    # Network / timeout errors (request may have been processed server-side)
    if (
        "timed out" in msg_lower
        or "readtimeout" in msg_lower
        or "timeout" in msg_lower
        or "connection" in msg_lower and "timeout" in msg_lower
    ):
        return (
            "Registration request timed out. If you received a confirmation email, your registration was successful. "
            "Otherwise, please try again."
        )

    # Never return raw/sanitized backend errors to users.
    if raw_msg and not UNIT_TESTING:
        app.logger.warning("Supabase sign-up failed (sanitized): %s", raw_msg)
    return "Registration failed. Please try again later."


def _is_transient_signup_error(err: Exception) -> bool:
    """Return True for transient/network sign-up failures.

    Supabase (GoTrue) can still create the user and send the confirmation email even if
    the client times out waiting for the response.
    """
    msg = _sanitize_public_error_message(str(err)).lower()
    return (
        "timed out" in msg
        or "readtimeout" in msg
        or "timeout" in msg
        or ("connection" in msg and "timeout" in msg)
    )


def _is_transient_auth_email_error(err: Exception) -> bool:
    """Return True for transient/network errors while requesting an auth email.

    Supabase can still send password reset emails even if the client times out.
    When we hit these errors, we should show a success-like message so users
    aren't confused when the email arrives.
    """
    msg = _sanitize_public_error_message(str(err)).lower()
    return (
        "timed out" in msg
        or "readtimeout" in msg
        or "timeout" in msg
        or "connection" in msg
        or "connection reset" in msg
        or "connection aborted" in msg
        or "temporarily unavailable" in msg
        or "service unavailable" in msg
        or "bad gateway" in msg
        or "gateway timeout" in msg
        or "502" in msg
        or "503" in msg
        or "504" in msg
    )


def _safe_predict(m, X):
    """Predict with XGBoost models across versions.

    Some XGBoost versions enforce feature-name validation even when given a
    pandas DataFrame. Disabling validate_features avoids Runtime 500s on Render.
    """
    try:
        return m.predict(X, validate_features=False)
    except TypeError:
        return m.predict(X)


def _sanitize_next_path(next_url: str | None) -> str | None:
    """Return a safe internal redirect target.

    Only allows relative paths (e.g. '/home') or same-host absolute URLs for APP_BASE_URL.
    Returns None for unsafe targets.
    """
    if not next_url:
        return None
    candidate = str(next_url).strip()
    if not candidate:
        return None

    parsed = urlparse(candidate)
    if not parsed.scheme and not parsed.netloc:
        return candidate if candidate.startswith("/") else None

    try:
        base = urlparse(APP_BASE_URL)
    except Exception:
        return None

    if parsed.scheme == base.scheme and parsed.netloc == base.netloc:
        path = parsed.path or "/"
        if parsed.query:
            path = path + "?" + parsed.query
        if parsed.fragment:
            path = path + "#" + parsed.fragment
        return path
    return None


def _issue_form_nonce(session_key: str) -> str:
    """Issue and store a one-time form nonce in the user session."""
    nonce = uuid4().hex
    session[session_key] = nonce
    return nonce


def _read_dataset_file(path: str, ext: str) -> pd.DataFrame:
    """Read uploaded dataset into a DataFrame based on extension."""
    if ext == ".csv":
        return pd.read_csv(path)
    if ext == ".xls":
        return pd.read_excel(path, engine="xlrd")
    return pd.read_excel(path)


def _load_report_from_session() -> dict:
    report_id = session.get("report_id")
    if not report_id:
        raise RuntimeError("No report found in session. Upload a dataset first.")

    report_path = os.path.join(DOWNLOAD_DIR, f"{report_id}.json")
    with open(report_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _cleanup_runtime_dirs(max_age_hours: float | None = None) -> None:
    """Best-effort cleanup of temp files to prevent disk growth on Render."""
    if max_age_hours is None:
        max_age_hours = CONFIG.cleanup_max_age_hours
    max_age_seconds = float(max_age_hours) * 3600.0

    now = time.time()
    for directory in (UPLOAD_DIR, GRAPH_DIR, DOWNLOAD_DIR):
        try:
            for name in os.listdir(directory):
                path = os.path.join(directory, name)
                try:
                    if not os.path.isfile(path):
                        continue
                    age = now - os.path.getmtime(path)
                    if age > max_age_seconds:
                        os.remove(path)
                except OSError:
                    # Cleanup failures should not affect user-facing responses.
                    app.logger.warning("Cleanup: failed removing file: %s", path)
                    continue
        except OSError:
            app.logger.warning("Cleanup: failed listing dir: %s", directory)
            continue


def _trend_label(series: Union[pd.Series, np.ndarray, List[float]], eps: float | None = None) -> str:
    """Return a human-friendly trend label based on the fitted slope."""
    if eps is None:
        eps = CONFIG.trend_eps
    values = np.asarray(series, dtype=float)
    values = values[np.isfinite(values)]
    if values.size < 2:
        return "Insufficient Data"

    x = np.arange(values.size, dtype=float)
    slope = np.polyfit(x, values, 1)[0]
    eps = float(eps)
    if slope > eps:
        return "Upward"
    if slope < -eps:
        return "Downward"
    return "Flat"


# ------------------------------
# PDF Font caching / idempotent registration
# ------------------------------

_PDF_FONT_LOCK = threading.Lock()
_PDF_FONT_NAMES: tuple[str, str] | None = None

_FONT_CACHE_LOCKS_GUARD = threading.Lock()
_FONT_CACHE_LOCKS: dict[str, threading.Lock] = {}


def _get_path_lock(path: str) -> threading.Lock:
    """Return a stable per-path lock for in-process concurrency control."""
    with _FONT_CACHE_LOCKS_GUARD:
        lock = _FONT_CACHE_LOCKS.get(path)
        if lock is None:
            lock = threading.Lock()
            _FONT_CACHE_LOCKS[path] = lock
        return lock


def _looks_like_font_bytes(data: bytes) -> bool:
    """Basic validation to avoid caching HTML error pages as fonts."""
    if not data or len(data) < 12:
        return False
    head = data[:4]
    # TrueType: 0x00010000, OpenType: 'OTTO', TTC: 'ttcf'
    return head in (b"\x00\x01\x00\x00", b"OTTO", b"ttcf")


def _atomic_write_bytes(dest_path: str, data: bytes) -> None:
    dest_dir = os.path.dirname(dest_path) or "."
    os.makedirs(dest_dir, exist_ok=True)
    tmp_path = dest_path + f".{uuid4().hex}.tmp"
    try:
        with open(tmp_path, "wb") as f:
            f.write(data)
            f.flush()
            os.fsync(f.fileno())

        # Atomic replace prevents corrupted cache state under concurrent writers.
        try:
            os.replace(tmp_path, dest_path)
        except OSError as replace_err:
            # Cross-device / filesystem fallback.
            try:
                shutil.copyfile(tmp_path, dest_path)
            except Exception:
                raise CacheWriteError(f"Failed writing cache file: {dest_path}") from replace_err
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass


def _font_file_is_valid(path: str) -> bool:
    try:
        with open(path, "rb") as f:
            head = f.read(4)
        return head in (b"\x00\x01\x00\x00", b"OTTO", b"ttcf")
    except OSError:
        return False


def _ensure_cached_font(url: str, dest_path: str, timeout_s: int = 20) -> bool:
    """Ensure a font file exists at dest_path.

    Uses atomic rename to avoid corruption under concurrency.
    Performs simple signature validation; registration will further validate.
    """
    try:
        with _get_path_lock(dest_path):
            if os.path.exists(dest_path) and _font_file_is_valid(dest_path):
                return True

            import urllib.request

            req = urllib.request.Request(
                url,
                headers={"User-Agent": "RetailDemandForecasting/1.0"},
            )
            with urllib.request.urlopen(req, timeout=timeout_s) as resp:
                data = resp.read()
            if not _looks_like_font_bytes(data):
                return False
            _atomic_write_bytes(dest_path, data)
            return _font_file_is_valid(dest_path)
    except Exception:
        app.logger.exception("Failed ensuring cached font: %s", url)
        return False


def _register_ttf_if_needed(font_name: str, font_path: str) -> bool:
    """Register a TTF/OTF font idempotently (safe to call multiple times)."""
    try:
        if not font_path or not os.path.exists(font_path):
            return False
        if font_name in pdfmetrics.getRegisteredFontNames():
            return True
        pdfmetrics.registerFont(TTFont(font_name, font_path))
        return True
    except Exception:
        app.logger.exception("Failed registering TTF font: %s", font_path)
        return False


def _get_pdf_font_names() -> tuple[str, str]:
    """Return (regular_font_name, bold_font_name) for PDF rendering.

    - Caches discovery results per-process to avoid repeated filesystem scans.
    - Ensures consistent regular/bold pairing.
    - Safe under concurrent requests.
    """
    global _PDF_FONT_NAMES
    with _PDF_FONT_LOCK:
        if _PDF_FONT_NAMES is not None:
            return _PDF_FONT_NAMES

        font_regular = "Helvetica"
        font_bold = "Helvetica-Bold"

        base_dir = os.path.dirname(os.path.abspath(__file__))

        repo_regular = [
            os.path.join(base_dir, "static", "fonts", "DejaVuSans.ttf"),
            os.path.join(base_dir, "static", "fonts", "NotoSans-Regular.ttf"),
            os.path.join(base_dir, "fonts", "DejaVuSans.ttf"),
            os.path.join(base_dir, "fonts", "NotoSans-Regular.ttf"),
        ]
        repo_bold = [
            os.path.join(base_dir, "static", "fonts", "DejaVuSans-Bold.ttf"),
            os.path.join(base_dir, "static", "fonts", "NotoSans-Bold.ttf"),
            os.path.join(base_dir, "fonts", "DejaVuSans-Bold.ttf"),
            os.path.join(base_dir, "fonts", "NotoSans-Bold.ttf"),
        ]

        linux_regular = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        ]
        linux_bold = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
        ]

        windir = os.environ.get("WINDIR", r"C:\\Windows")
        fonts_dir = os.path.join(windir, "Fonts")
        windows_regular = [
            os.path.join(fonts_dir, "segoeui.ttf"),
            os.path.join(fonts_dir, "seguisym.ttf"),
            os.path.join(fonts_dir, "arial.ttf"),
        ]
        windows_bold = [
            os.path.join(fonts_dir, "segoeuib.ttf"),
            os.path.join(fonts_dir, "arialbd.ttf"),
        ]

        regular_candidates = repo_regular + linux_regular + windows_regular
        bold_candidates = repo_bold + linux_bold + windows_bold

        # Prefer system/repo fonts first (no network dependency)
        if any(_register_ttf_if_needed("PDFUnicode", p) for p in regular_candidates):
            font_regular = "PDFUnicode"
        if any(_register_ttf_if_needed("PDFUnicode-Bold", p) for p in bold_candidates):
            font_bold = "PDFUnicode-Bold"

        # If bold isn't available but regular is, keep pairing consistent.
        if font_regular != "Helvetica" and font_bold in ("Helvetica-Bold", "Helvetica"):
            font_bold = font_regular

        # Final fallback: cache-download DejaVu into runtime dir.
        if font_regular == "Helvetica":
            fonts_cache_dir = os.path.join(RUNTIME_DIR, "pdf_fonts")
            dejavu_regular_path = os.path.join(fonts_cache_dir, "DejaVuSans.ttf")
            dejavu_bold_path = os.path.join(fonts_cache_dir, "DejaVuSans-Bold.ttf")
            dejavu_regular_url = "https://raw.githubusercontent.com/dejavu-fonts/dejavu-fonts/version_2_37/ttf/DejaVuSans.ttf"
            dejavu_bold_url = "https://raw.githubusercontent.com/dejavu-fonts/dejavu-fonts/version_2_37/ttf/DejaVuSans-Bold.ttf"

            if _ensure_cached_font(dejavu_regular_url, dejavu_regular_path) and _register_ttf_if_needed(
                "PDFUnicode", dejavu_regular_path
            ):
                font_regular = "PDFUnicode"
            if _ensure_cached_font(dejavu_bold_url, dejavu_bold_path) and _register_ttf_if_needed(
                "PDFUnicode-Bold", dejavu_bold_path
            ):
                font_bold = "PDFUnicode-Bold"

            if font_regular != "Helvetica" and font_bold in ("Helvetica-Bold", "Helvetica"):
                font_bold = font_regular

        _PDF_FONT_NAMES = (font_regular, font_bold)
        return _PDF_FONT_NAMES

# ---------- LANDING / LOGIN ----------
@app.route('/', methods=['GET', 'POST'])
def landing():
    # Public homepage (always shown as the default page).
    # Backward compatibility: if an older login form posts to '/', handle it like '/login'.
    if request.method == 'POST':
        return login()
    return render_template('homepage.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Supabase Auth expects an email for sign-in.
        email_raw = _first_non_empty(request.form.get('username'), request.form.get('email'))
        pwd = request.form.get('password')
        if not email_raw or not pwd:
            return render_template("login.html", error="Please enter email and password.")
        email = email_raw.lower()

        try:
            sb = get_supabase()
            res = sb.auth.sign_in_with_password({"email": email, "password": pwd})
        except Exception as e:
            # Avoid leaking internal configuration or stack traces to users.
            app.logger.exception("Supabase sign-in failed")
            msg_lower = str(e).lower()
            if "email not confirmed" in msg_lower or "email_not_confirmed" in msg_lower:
                return render_template("login.html", error="Email not confirmed. Please confirm from your inbox, then login.")
            if "invalid login credentials" in msg_lower or "invalid_grant" in msg_lower:
                return render_template("login.html", error="Invalid email or password.")
            if "missing supabase config" in msg_lower:
                return render_template("login.html", error="Login is temporarily unavailable due to a server configuration issue.")
            return render_template("login.html", error="Login failed. Please try again later.")

        user_obj = _resp_get(res, "user")
        session_obj = _resp_get(res, "session")

        if user_obj is None:
            return render_template("login.html", error="Invalid email or password.")

        # If Confirm Email is enabled, Supabase can return a user but no session.
        if session_obj is None:
            return render_template(
                "login.html",
                error="Login blocked. Please confirm your email first, then login."
            )

        access_token = _resp_get(session_obj, "access_token")
        refresh_token = _resp_get(session_obj, "refresh_token")
        if not access_token:
            return render_template("login.html", error="Login failed. Please try again later.")

        # Avoid stale data by clearing any previous session before setting new values.
        session.clear()
        session['logged_in'] = True
        session['user_id'] = _resp_get(user_obj, "id")
        session['user_email'] = _resp_get(user_obj, "email")
        session['email_verified'] = _compute_email_verified_from_user_obj(user_obj)
        session['access_token'] = access_token
        session['refresh_token'] = refresh_token

        # Best-effort: fetch profile (requires your RLS policies + access token)
        try:
            sb_user = get_supabase(session['access_token'])
            profile_res = (
                sb_user.table("profiles")
                .select("id, username")
                .eq("id", session['user_id'])
                .maybe_single()
                .execute()
            )
            if getattr(profile_res, "data", None):
                session['username'] = profile_res.data.get('username')
        except Exception:
            # Don't block login if profile fetch fails, but don't fail silently.
            app.logger.warning("Profile fetch failed for user_id=%s", session.get('user_id'), exc_info=True)

        return redirect(url_for('home'))
    return render_template("login.html")

# ---------- REGISTER ----------
@app.route('/register', methods=['GET', 'POST'])
def register():
    policy = _password_policy_payload()
    nonce_key = "register_nonce"
    if request.method == 'POST':
        form_nonce = (request.form.get(nonce_key) or "").strip()
        expected_nonce = (session.pop(nonce_key, None) or "").strip()
        if not form_nonce or not expected_nonce or form_nonce != expected_nonce:
            return render_template(
                "register.html",
                error=(
                    "Registration already submitted. If you received a confirmation email, please confirm it, then login."
                ),
                password_policy=policy,
                register_nonce=_issue_form_nonce(nonce_key),
            )

        # Supabase Auth expects an email for sign-up.
        email_raw = _first_non_empty(request.form.get('username'), request.form.get('email'))
        if not email_raw:
            return render_template(
                "register.html",
                error="Please enter your email.",
                password_policy=policy,
                register_nonce=_issue_form_nonce(nonce_key),
            )
        email = email_raw.lower()
        # Optional username field (if not provided, derive from email prefix)
        username = (request.form.get('profile_username') or "").strip()
        if not username and '@' in email:
            username = email.split('@', 1)[0]

        pwd = request.form.get('password')
        confirm = request.form.get('confirm') or request.form.get('confirm_password')
        if not pwd or not confirm:
            return render_template(
                "register.html",
                error="Please enter and confirm your password.",
                password_policy=policy,
                register_nonce=_issue_form_nonce(nonce_key),
            )

        if pwd != confirm:
            return render_template(
                "register.html",
                error="Passwords do not match.",
                password_policy=policy,
                register_nonce=_issue_form_nonce(nonce_key),
            )

        strength_err = _password_strength_error(pwd)
        if strength_err:
            return render_template(
                "register.html",
                error=strength_err,
                password_policy=policy,
                register_nonce=_issue_form_nonce(nonce_key),
            )

        try:
            sb = get_supabase()

            # Never construct redirect URLs from request headers (open redirect risk).
            redirect_base = CONFIG.email_redirect_base_url

            # Send username in user metadata so your DB trigger can populate public.profiles.username
            res = sb.auth.sign_up(
                {
                    "email": email,
                    "password": pwd,
                    "options": {
                        "data": {"username": username},
                        # Ensure the confirmation link redirects back to this Flask app
                        "emailRedirectTo": f"{redirect_base}/auth/callback",
                    },
                }
            )
        except Exception as e:
            # A network timeout can occur even if Supabase already created the user and
            # sent the confirmation email. Treat it as an indeterminate-but-likely-success
            # outcome with clear next steps.
            if _is_transient_signup_error(e):
                if not UNIT_TESTING:
                    app.logger.warning("Supabase sign-up timed out: %s", _sanitize_public_error_message(str(e)))
                return render_template(
                    "register.html",
                    success=(
                        "If you received a confirmation email, your registration was successful. "
                        "Please confirm your email, then login."
                    ),
                    password_policy=policy,
                    form_disabled=True,
                )

            app.logger.exception("Supabase sign-up failed")
            return render_template(
                "register.html",
                error=_friendly_register_error(e),
                password_policy=policy,
                register_nonce=_issue_form_nonce(nonce_key),
            )

        user_obj = _resp_get(res, "user")
        session_obj = _resp_get(res, "session")

        if user_obj is None:
            return render_template(
                "register.html",
                error="Registration failed. Please try again later.",
                password_policy=policy,
                register_nonce=_issue_form_nonce(nonce_key),
            )

        # If Supabase returned an existing-user-shaped response, show a clear message.
        if _looks_like_existing_user_after_signup(user_obj):
            return render_template(
                "register.html",
                error="This email is already registered. Please login or reset your password.",
                password_policy=policy,
                register_nonce=_issue_form_nonce(nonce_key),
            )

        # With email confirmation enabled, session may be None: that's still a success.
        # If email confirmation is disabled, Supabase may return a session; we can auto-login.
        access_token = _resp_get(session_obj, "access_token")
        refresh_token = _resp_get(session_obj, "refresh_token")
        if access_token:
            session.clear()
            session['logged_in'] = True
            session['user_id'] = _resp_get(user_obj, "id")
            session['user_email'] = _resp_get(user_obj, "email")
            session['email_verified'] = _compute_email_verified_from_user_obj(user_obj)
            session['access_token'] = access_token
            session['refresh_token'] = refresh_token
            return redirect(url_for('home'))

        return render_template(
            "register.html",
            success="Registration successful. Check your email to confirm, then login.",
            password_policy=policy,
            form_disabled=True,
        )

    return render_template(
        "register.html",
        password_policy=policy,
        register_nonce=_issue_form_nonce(nonce_key),
    )


@app.route('/auth/callback')
def auth_callback():
    """Landing page for Supabase email confirmation redirect.

    Supabase handles the verification itself and then redirects here.
    We just show a friendly message and send the user to login.
    """
    err = request.args.get('error') or request.args.get('error_code')
    err_desc = request.args.get('error_description')
    if err or err_desc:
        raw_msg = _sanitize_public_error_message(err_desc or err)
        if raw_msg and not UNIT_TESTING:
            app.logger.warning("Auth callback error (sanitized): %s", raw_msg)
        return render_template("login.html", error="Email confirmation failed. Please try again.")

    return render_template(
        "login.html",
        error=None,
        info="Email confirmed. You can login now."
    )


# ---------- FORGOT / RESET PASSWORD ----------
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = (request.form.get('email') or "").strip().lower()
        if not email:
            return render_template("forgot_password.html", error="Please enter your email.")

        try:
            sb = get_supabase()
            # Never construct redirect URLs from request headers (open redirect risk).
            redirect_base = CONFIG.email_redirect_base_url
            redirect_to = f"{redirect_base}/reset-password"

            # supabase-py has had slightly different method signatures across versions
            try:
                sb.auth.reset_password_email(email, {"redirect_to": redirect_to})
            except TypeError:
                sb.auth.reset_password_email(email)
            except AttributeError:
                # older method name
                sb.auth.reset_password_for_email(email, {"redirect_to": redirect_to})

            return render_template(
                "forgot_password.html",
                info="If an account exists for that email, a password reset link has been sent."
            )
        except Exception as e:
            app.logger.exception("Supabase reset password email failed")
            raw_msg = _sanitize_public_error_message(str(e))
            msg_lower = raw_msg.lower()

            # Misconfiguration should be shown as a clear failure.
            if "missing supabase config" in msg_lower:
                return render_template(
                    "forgot_password.html",
                    error="Password reset is temporarily unavailable due to a server configuration issue.",
                )

            # Rate limiting / abuse protection.
            if (
                "rate limit" in msg_lower
                or "too many requests" in msg_lower
                or "over_email_send_rate_limit" in msg_lower
                or "429" in msg_lower
            ):
                return render_template(
                    "forgot_password.html",
                    error="Too many attempts. Please wait and try again later.",
                )

            # If the request may have succeeded server-side (timeouts / transient network),
            # show a success-like message to avoid confusing users who still receive the email.
            if _is_transient_auth_email_error(e):
                if not UNIT_TESTING:
                    app.logger.warning("Reset password request may have timed out (sanitized): %s", raw_msg)
                return render_template(
                    "forgot_password.html",
                    info=(
                        "If an account exists for that email, a password reset link has been sent. "
                        "Please check your inbox (and spam) in a minute."
                    ),
                )

            # Default: real error.
            return render_template(
                "forgot_password.html",
                error="Failed to send reset email. Please try again later.",
            )

    return render_template("forgot_password.html")


@app.route('/reset-password', methods=['GET'])
def reset_password():
    # The Supabase recovery link typically includes tokens in the URL fragment (#...)
    # which the server cannot read. The template uses JS to read the fragment and
    # complete the password update via supabase-js.
    supabase_url = CONFIG.supabase_url
    supabase_anon_key = CONFIG.supabase_anon_key
    if not supabase_url or not supabase_anon_key:
        if not UNIT_TESTING:
            app.logger.warning("Reset password page requested but Supabase env vars are missing")
        return render_template(
            "reset_password.html",
            error="Password reset is not configured on this server.",
            supabase_url="",
            supabase_anon_key="",
        )

    # Optional query-param tokens (some auth flows may use query instead of hash)
    access_q = (request.args.get("access_token") or "").strip()
    refresh_q = (request.args.get("refresh_token") or "").strip()
    try:
        if (access_q or refresh_q) and not (access_q and refresh_q):
            raise InvalidResetTokens("missing_query_tokens")
    except InvalidResetTokens:
        return render_template(
            "reset_password.html",
            error="Missing recovery tokens. Please open this page from the reset email link.",
            supabase_url=supabase_url,
            supabase_anon_key=supabase_anon_key,
        )

    return render_template(
        "reset_password.html",
        supabase_url=supabase_url,
        supabase_anon_key=supabase_anon_key,
    )

# ---------- HOME ----------
@app.route('/home')
def home():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return render_template("index.html")


# ---------- SETTINGS ----------
@app.route('/settings', methods=['GET'])
def settings_page():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    # Keep the session flag reasonably fresh.
    _refresh_sender_verified_state_best_effort()

    return render_template("settings.html")

# ---------- LOGOUT ----------
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ---------- UPLOAD & FORECAST ----------
@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    if request.method == 'GET':
        return redirect(url_for('home'))

    app.logger.info("/upload POST received")

    try:
        _cleanup_runtime_dirs()

        file = request.files.get('file')
        if not file or not file.filename:
            return render_template("index.html", error="Please choose a dataset file to upload."), 400

        months_raw = (request.form.get('months') or "").strip()
        if not months_raw.isdigit():
            return render_template("index.html", error="Please enter a valid number of months."), 400
        months = int(months_raw)
        if months < 1 or months > MAX_FORECAST_MONTHS:
            return render_template(
                "index.html",
                error=f"Months must be between 1 and {MAX_FORECAST_MONTHS}.",
            ), 400

        filename = secure_filename(file.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in ALLOWED_UPLOAD_EXTS:
            return render_template(
                "index.html",
                error=(
                    f"Unsupported file type: {ext or '(no extension)'}. "
                    f"Allowed: {', '.join(sorted(ALLOWED_UPLOAD_EXTS))}."
                ),
            ), 400
        report_id = str(uuid4())

        dataset_path = os.path.join(UPLOAD_DIR, f"{report_id}{ext or '.xlsx'}")
        file.save(dataset_path)

        try:
            df = _read_dataset_file(dataset_path, ext)
        except Exception:
            app.logger.exception("Failed reading uploaded file: %s", filename)
            return render_template(
                "index.html",
                error="Unable to read the uploaded file. Please upload a valid .csv or Excel file.",
            ), 400

        df.columns = df.columns.str.lower().str.strip()

        required_cols = {"date", "units_sold"}
        missing = sorted(required_cols - set(df.columns))
        if missing:
            return render_template(
                "index.html",
                error=f"Missing required columns: {', '.join(missing)}. Required: date, units_sold.",
            ), 400

        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df['units_sold'] = pd.to_numeric(df['units_sold'], errors='coerce')
        df = df.sort_values('date').dropna(subset=['date', 'units_sold'])

        raw_history = df['units_sold'].astype(float).tolist()
        if len(raw_history) < 30:
            return render_template(
                "index.html",
                error="Not enough history for forecasting. Please upload at least 30 days of data.",
            ), 400

        # ---------- FEATURE ENGINEERING ----------
        df_feat = df.copy()
        df_feat['lag_1'] = df_feat['units_sold'].shift(1)
        df_feat['lag_7'] = df_feat['units_sold'].shift(7)
        df_feat['roll_mean_7'] = df_feat['units_sold'].shift(1).rolling(7).mean()
        df_feat['roll_mean_30'] = df_feat['units_sold'].shift(1).rolling(30).mean()
        df_feat.dropna(inplace=True)

        if df_feat.empty:
            return render_template(
                "index.html",
                error=(
                    "Not enough rows after feature engineering. "
                    "Please upload a dataset with at least 30+ days of sales history."
                ),
            ), 400

        X = df_feat[FEATURE_COLS].astype(float)
        y = df_feat['units_sold'].astype(float)
        preds = _safe_predict(model, X)

        # ---------- METRICS ----------
        mae = int(mean_absolute_error(y, preds))
        rmse = int(np.sqrt(mean_squared_error(y, preds)))
        r2 = round(r2_score(y, preds), 2)

        # ---------- PAST GRAPH ----------
        graph_df = pd.DataFrame({
            'date': df_feat['date'],
            'actual': y,
            'predicted': preds
        }).set_index('date').resample('W').mean()

        past_actual_trend = _trend_label(graph_df['actual'])
        past_pred_trend = _trend_label(graph_df['predicted'])

        past_image = f"past_{report_id}.png"
        future_image = f"future_{report_id}.png"
        past_path = os.path.join(GRAPH_DIR, past_image)
        future_path = os.path.join(GRAPH_DIR, future_image)

        plt.figure(figsize=(11, 5))
        plt.plot(graph_df.index, graph_df['actual'], label="Actual", linewidth=2)
        plt.plot(graph_df.index, graph_df['predicted'], label="Predicted", linewidth=2)
        plt.legend()
        plt.grid(True)
        plt.title("Actual vs Predicted Sales (Weekly)")
        plt.xlabel("Date")
        plt.ylabel("Units Sold")
        # Show month names only (no year) on the x-axis
        ax = plt.gca()
        ax.xaxis.set_major_locator(mdates.MonthLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%b'))
        plt.xticks(rotation=0)
        plt.tight_layout()
        plt.savefig(past_path)
        plt.close()

        # ---------- FUTURE FORECAST ----------
        future: List[float] = []
        history = list(raw_history)

        steps = months * 4
        for _ in range(steps):
            lag_1 = history[-1]
            lag_7 = history[-7]
            roll_mean_7 = float(np.mean(history[-7:]))
            roll_mean_30 = float(np.mean(history[-30:]))

            # XGBoost validates feature names when the model was trained with them.
            # Use a DataFrame with the same column names to avoid ValueError on Render.
            features_df = pd.DataFrame(
                [[lag_1, lag_7, roll_mean_7, roll_mean_30]],
                columns=FEATURE_COLS,
            )
            val = float(_safe_predict(model, features_df)[0])
            future.append(val)
            history.append(val)

        monthly_pred = (
            pd.Series(future)
            .groupby(np.arange(len(future)) // 4)
            .mean()
        )

        future_df = pd.DataFrame({
            "Month": list(range(1, months + 1)),
            "Predicted_Sales": monthly_pred.round().astype(int).values,
        })

        future_forecast_table = future_df.to_dict(orient='records')

        future_trend = _trend_label(monthly_pred.values)

        plt.figure(figsize=(10, 5))
        plt.plot(future_df['Month'], future_df['Predicted_Sales'], marker='o')
        plt.title(f"Future Sales Forecast ({months} Month{'s' if months > 1 else ''})")
        plt.xlabel("Month")
        plt.ylabel("Predicted Sales (Units)")
        plt.xticks(future_df['Month'].astype(int).tolist())
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(future_path)
        plt.close()

        # ---------- INSIGHTS ----------
        avg_demand = int(future_df['Predicted_Sales'].mean())
        recommendation = (
            "High demand expected. Maintain high inventory."
            if avg_demand > 70 else
            "Moderate demand expected. Maintain balanced inventory."
            if avg_demand > 40 else
            "Low demand expected. Avoid overstocking."
        )

        # ---------- OPTIONAL CATEGORY COST TABLE (FUTURE FORECAST MONTHS) ----------
        # Requirement: show forecast months (1..N), NOT recent historical months, and month as numbers only.
        category_cost_table = []
        if 'price' in df.columns and 'product_category' in df.columns:
            df_cat = df[['date', 'product_category', 'units_sold', 'price']].copy()
            df_cat['date'] = pd.to_datetime(df_cat['date'], errors='coerce')
            df_cat['price'] = pd.to_numeric(df_cat['price'], errors='coerce')
            df_cat['units_sold'] = pd.to_numeric(df_cat['units_sold'], errors='coerce')
            df_cat = df_cat.dropna(subset=['date', 'product_category', 'units_sold', 'price'])

            if not df_cat.empty:
                # Build historical monthly totals per category
                df_cat['year_month'] = df_cat['date'].dt.to_period('M')
                df_cat['total_cost'] = df_cat['price'] * df_cat['units_sold']

                monthly_cat = (
                    df_cat.groupby(['year_month', 'product_category'])
                    .agg(
                        units=('units_sold', 'sum'),
                        cost=('total_cost', 'sum'),
                    )
                    .reset_index()
                )

                # Use the most recent months available (up to the requested forecast horizon)
                month_keys = sorted(monthly_cat['year_month'].unique())
                recent_months = month_keys[-months:] if month_keys else []
                if recent_months:
                    recent = monthly_cat[monthly_cat['year_month'].isin(recent_months)].copy()
                else:
                    recent = monthly_cat.copy()

                if not recent.empty:
                    # Average monthly units per category from the uploaded dataset
                    avg_units_by_cat = recent.groupby('product_category')['units'].mean()
                    baseline_total = float(avg_units_by_cat.sum())

                    # Units-weighted average price per category from the uploaded dataset
                    price_weighted = df_cat.groupby('product_category').apply(
                        lambda g: (g['price'] * g['units_sold']).sum() / max(g['units_sold'].sum(), 1.0)
                    )
                    price_weighted = price_weighted.replace([np.inf, -np.inf], np.nan).fillna(0.0)

                    if baseline_total > 0:
                        # Scale using the forecast's *relative* month-to-month changes,
                        # but keep the magnitude based on the uploaded dataset.
                        forecast_vals = future_df['Predicted_Sales'].astype(float).values
                        denom = float(np.mean(forecast_vals)) if forecast_vals.size else 1.0
                        denom = denom if denom != 0 else 1.0

                        categories = avg_units_by_cat.index.tolist()

                        for i in range(1, months + 1):
                            rel = float(forecast_vals[i - 1]) / denom
                            target_total = baseline_total * rel

                            # Allocate to categories proportionally to historical averages
                            remaining = target_total
                            for idx, cat in enumerate(categories):
                                if idx == len(categories) - 1:
                                    units = max(0, int(round(remaining)))
                                else:
                                    share = float(avg_units_by_cat.loc[cat]) / baseline_total
                                    units = max(0, int(round(target_total * share)))
                                    remaining -= units

                                avg_price = float(price_weighted.get(cat, 0.0))
                                total_cost = int(round(units * avg_price))
                                category_cost_table.append(
                                    {
                                        'month': i,
                                        'product_category': str(cat),
                                        'total_units_sold': int(units),
                                        'total_cost': int(total_cost),
                                    }
                                )

        # ---------- STORE FOR DOWNLOAD (avoid huge cookies) ----------
        report = {
            "months": months,
            "metrics": {'MAE': mae, 'RMSE': rmse, 'R2': r2},
            "insights": {'Average Demand': avg_demand, 'Recommendation': recommendation},
            "trends": {
                'Past Actual Trend': past_actual_trend,
                'Past Predicted Trend': past_pred_trend,
                'Future Forecast Trend': future_trend,
            },
            "future_forecast_table": future_forecast_table,
            "category_cost_table": category_cost_table,
            "past_image": past_image,
            "future_image": future_image,
        }

        report_path = os.path.join(DOWNLOAD_DIR, f"{report_id}.json")
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(report, f)

        session['report_id'] = report_id

        return render_template(
            "result.html",
            past_image=past_image,
            future_image=future_image,
            months=months,
            mae=mae,
            rmse=rmse,
            r2=r2,
            avg_demand=avg_demand,
            recommendation=recommendation,
            past_actual_trend=past_actual_trend,
            past_pred_trend=past_pred_trend,
            future_trend=future_trend,
            future_forecast_table=future_forecast_table,
            category_cost_table=category_cost_table
        )
    except Exception:
        app.logger.exception("Prediction failed in /upload")
        return render_template("index.html", error="Prediction failed. Please check your dataset format."), 500

# ---------- DOWNLOAD EXCEL ----------
@app.route('/download/excel')
def download_excel():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    try:
        report = _load_report_from_session()
    except (RuntimeError, FileNotFoundError, json.JSONDecodeError):
        return render_template("index.html", error="Report not found. Please upload a dataset again."), 400

    report_id = session.get("report_id")
    path = os.path.join(DOWNLOAD_DIR, f"retail_forecast_report_{report_id}.xlsx")

    _build_excel_report_file(report, path)

    response = send_file(path, as_attachment=True)

    @response.call_on_close
    def _remove_generated_excel() -> None:
        try:
            os.remove(path)
        except OSError:
            # Cleanup failures must not affect user response.
            app.logger.warning("Failed removing generated Excel: %s", path)

    return response

# ---------- DOWNLOAD PDF ----------
@app.route('/download/pdf')
def download_pdf():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    try:
        report = _load_report_from_session()
    except (RuntimeError, FileNotFoundError, json.JSONDecodeError):
        return render_template("index.html", error="Report not found. Please upload a dataset again."), 400

    report_id = session.get("report_id")
    path = os.path.join(DOWNLOAD_DIR, f"retail_forecast_report_{report_id}.pdf")
    _build_pdf_report_file(report, path)

    response = send_file(path, as_attachment=True)

    @response.call_on_close
    def _remove_generated_pdf() -> None:
        try:
            os.remove(path)
        except OSError:
            # Cleanup failures must not affect user response.
            app.logger.warning("Failed removing generated PDF: %s", path)

    return response


@app.route('/email/pdf', methods=['POST'])
def email_pdf():
    if not session.get('logged_in'):
        return jsonify({"ok": False, "message": "Please login first."}), 401

    email = None
    pdf_password = None
    try:
        if request.is_json:
            payload = request.get_json(silent=True) or {}
            email = (payload.get('email') or '').strip()
            pdf_password = payload.get("pdf_password") or payload.get("password")
        else:
            email = (request.form.get('email') or '').strip()
            pdf_password = request.form.get("pdf_password") or request.form.get("password")
    except Exception:
        email = None
        pdf_password = None

    email, email_err = _normalize_and_validate_email_for_sending(email)
    if email_err:
        return jsonify({"ok": False, "message": email_err}), 400

    # Optional abuse prevention: block disposable email domains.
    if BLOCK_DISPOSABLE_EMAILS and _is_disposable_email_domain(email):
        _audit_event_best_effort(
            "recipient_blocked_disposable",
            owner_id=(session.get("user_id") or None),
            actor_email=(session.get("user_email") or None),
            recipient_email=email,
            report_id=str(session.get("report_id") or "") or None,
            meta={},
        )
        return jsonify({"ok": False, "message": "Disposable email domains are not allowed."}), 400

    # If a password was provided, enforce strong requirements.
    pdf_password = (pdf_password or "").strip() if pdf_password is not None else ""
    if pdf_password:
        pw_err = _validate_pdf_password_or_error(pdf_password)
        if pw_err:
            return jsonify({"ok": False, "message": pw_err}), 400

    # Enforce sender verification before any delivery.
    sender, sender_err = _require_verified_sender()
    if sender_err:
        return jsonify(sender_err[0]), sender_err[1]

    report_id = _parse_report_id(session.get("report_id"))
    if not report_id:
        return jsonify({"ok": False, "message": "Report not found. Please upload a dataset again."}), 400

    # Rate limiting: limit report send attempts per sender.
    rl_sender_key = f"send_report:{sender.id}"
    if _rate_limit_hit(rl_sender_key, REPORT_SEND_RATE_LIMIT, REPORT_SEND_RATE_WINDOW_SECONDS):
        _audit_event_best_effort(
            "rate_limited_send",
            owner_id=sender.id,
            actor_email=sender.email,
            recipient_email=email,
            report_id=report_id,
            meta={"scope": "sender"},
        )
        return jsonify({"ok": False, "message": "Too many send attempts. Please try again later."}), 429

    # Anti-spam: cooldown per recipient.
    cd_recipient_key = f"send_to:{sender.id}:{email.lower().strip()}"
    if _cooldown_block(cd_recipient_key, RECIPIENT_COOLDOWN_SECONDS):
        _audit_event_best_effort(
            "cooldown_send",
            owner_id=sender.id,
            actor_email=sender.email,
            recipient_email=email,
            report_id=report_id,
            meta={"cooldown_seconds": RECIPIENT_COOLDOWN_SECONDS},
        )
        return jsonify({"ok": False, "message": "Please wait a moment before sending to the same recipient again."}), 429

    # Authorize delivery. If not allowed, initiate recipient verification (one-time share) and do NOT deliver.
    if not can_send_report(sender, email, report_id):
        cfg = _get_sendgrid_config()
        if not cfg.get("api_key") or not cfg.get("from_email"):
            return jsonify({
                "ok": False,
                "message": "Recipient email is not verified for this report.",
            }), 403

        try:
            access_token = (session.get("access_token") or "").strip()
            if not access_token:
                return jsonify({"ok": False, "message": "Login session expired. Please login again."}), 401

            _create_or_refresh_report_share_verification(sender.id, report_id, email, access_token, pdf_password=pdf_password or None)
            _audit_event_best_effort(
                "share_verification_sent",
                owner_id=sender.id,
                actor_email=sender.email,
                recipient_email=email,
                report_id=report_id,
                meta={"password_protected": bool(pdf_password)},
            )
            return jsonify({
                "ok": True,
                "status": "verification_sent",
                "message": "Verification email sent to recipient. Ask them to verify, then resend the report.",
            }), 202
        except Exception as e:
            msg = str(e).lower()
            if "missing_sendgrid_config" in msg:
                return jsonify({"ok": False, "message": "Email service is not configured on the server."}), 500
            if "rate_limited" in msg:
                return jsonify({"ok": False, "message": "Too many verification emails. Please try again later."}), 429
            if "cooldown" in msg:
                return jsonify({"ok": False, "message": "Please wait before sending another verification email to this recipient."}), 429
            if "password_protection_not_configured" in msg or "password_storage_unavailable" in msg:
                # Verification email was sent, but the password could not be stored for later.
                _audit_event_best_effort(
                    "share_verification_sent_password_not_stored",
                    owner_id=sender.id,
                    actor_email=sender.email,
                    recipient_email=email,
                    report_id=report_id,
                    meta={"password_protected_requested": bool(pdf_password)},
                )
                return jsonify({
                    "ok": True,
                    "status": "verification_sent",
                    "message": "Verification email sent to recipient. Password cannot be stored on this server, so you must re-enter the password when you resend after verification.",
                }), 202
            app.logger.exception("Failed initiating recipient verification")
            return jsonify({"ok": False, "message": "Failed to send verification email. Please try again later."}), 500

    # If the sender previously requested password protection for this share, require encryption.
    if not pdf_password and _share_password_was_requested(sender.id, report_id, email):
        stored_pw = _get_share_pdf_password(sender.id, report_id, email)
        if stored_pw:
            pdf_password = stored_pw
        else:
            return jsonify({
                "ok": False,
                "message": "Password for this share expired. Please set a new PDF password and resend.",
            }), 409

    return _deliver_report_pdf_via_email(email, report_id, sender=sender, pdf_password=pdf_password or None)


def _deliver_report_pdf_via_email(recipient_email: str, report_id: str, sender: CurrentUser | None = None, pdf_password: str | None = None) -> tuple:
    """Generate the current session report files (PDF+XLSX) and email them.

    Backward compatible name: route paths still reference /email/pdf.
    If a password is provided, we send 2 attachments:
    - password-protected PDF (PDF encryption)
    - password-protected Excel (password-to-open on Windows+Excel when available; otherwise edit-protected XLSX)
    """
    try:
        report = _load_report_from_session()
    except (RuntimeError, FileNotFoundError, json.JSONDecodeError):
        return jsonify({"ok": False, "message": "Report not found. Please upload a dataset again."}), 400

    pdf_path = os.path.join(DOWNLOAD_DIR, f"retail_forecast_report_{report_id}.pdf")
    xlsx_path = os.path.join(DOWNLOAD_DIR, f"retail_forecast_report_{report_id}.xlsx")
    filename_pdf = f"retail_forecast_report_{report_id}.pdf"
    filename_xlsx = f"retail_forecast_report_{report_id}.xlsx"

    try:
        _build_pdf_report_file(report, pdf_path)
        _build_excel_report_file(report, xlsx_path)
        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()
        with open(xlsx_path, 'rb') as f:
            xlsx_bytes = f.read()

        pdf_password = (pdf_password or "").strip() if pdf_password is not None else ""

        subject = "Demand Forecasting Report"

        if pdf_password:
            pdf_bytes = _encrypt_pdf_bytes_if_requested(pdf_bytes, pdf_password)

            # Prefer "password-to-open" XLSX on Windows when Excel is installed.
            xlsx_open_pw_bytes = _encrypt_xlsx_bytes_password_to_open_or_none(xlsx_bytes, pdf_password)
            if xlsx_open_pw_bytes:
                body = "Attached are your password-protected report files (PDF and Excel)."
                delivered_kind = "pdf+xlsx_openpw"
                attachments = [
                    (filename_pdf, "application/pdf", pdf_bytes),
                    (filename_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", xlsx_open_pw_bytes),
                ]
            else:
                # Fallback: OpenPyXL cannot do password-to-open encryption, but we can still
                # apply worksheet/workbook protection (edits require password).
                protected_xlsx_bytes = _protect_xlsx_bytes_if_requested(xlsx_bytes, pdf_password)

                body = "Attached are your password-protected report files (PDF and Excel)."
                delivered_kind = "pdf+xlsx_editpw"
                attachments = [
                    (filename_pdf, "application/pdf", pdf_bytes),
                    (filename_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", protected_xlsx_bytes),
                ]
        else:
            body = "Attached are your Retail Demand Forecasting report files (PDF and Excel)."
            delivered_kind = "pdf+xlsx"
            attachments = [
                (filename_pdf, "application/pdf", pdf_bytes),
                (filename_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", xlsx_bytes),
            ]

        _send_email_with_attachments_sendgrid(recipient_email, subject, body, attachments)

        _audit_event_best_effort(
            "report_delivered",
            owner_id=(sender.id if sender else (session.get("user_id") or None)),
            actor_email=(sender.email if sender else (session.get("user_email") or None)),
            recipient_email=recipient_email,
            report_id=report_id,
            meta={"password_protected": bool(pdf_password), "delivered": delivered_kind},
        )
        return jsonify({"ok": True, "message": "Report sent successfully."}), 200
    except Exception as e:
        msg = str(e).lower()
        if "missing_sendgrid_config" in msg:
            return jsonify({"ok": False, "message": "Email service is not configured on the server."}), 500
        if "pdf_encryption_unavailable" in msg:
            return jsonify({"ok": False, "message": "PDF encryption is not available on the server."}), 503
        if "zip_encryption_unavailable" in msg or "xlsx_protection_unavailable" in msg:
            return jsonify({"ok": False, "message": "Excel password protection is not available on the server."}), 503
        app.logger.exception("Failed sending report email")
        _audit_event_best_effort(
            "report_delivery_failed",
            owner_id=(sender.id if sender else (session.get("user_id") or None)),
            actor_email=(sender.email if sender else (session.get("user_email") or None)),
            recipient_email=recipient_email,
            report_id=report_id,
            meta={},
        )
        return jsonify({"ok": False, "message": "Failed to send email. Please try again later."}), 500
    finally:
        try:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)
        except OSError:
            app.logger.warning("Failed removing emailed report files: %s", report_id)


@app.route('/email/pdf/self', methods=['POST'])
def email_pdf_self():
    """Default delivery method: send report only to the sender's verified email."""
    sender, sender_err = _require_verified_sender()
    if sender_err:
        return jsonify(sender_err[0]), sender_err[1]

    pdf_password = None
    try:
        if request.is_json:
            payload = request.get_json(silent=True) or {}
            pdf_password = payload.get("pdf_password") or payload.get("password")
        else:
            pdf_password = request.form.get("pdf_password") or request.form.get("password")
    except Exception:
        pdf_password = None

    pdf_password = (pdf_password or "").strip() if pdf_password is not None else ""
    if pdf_password:
        pw_err = _validate_pdf_password_or_error(pdf_password)
        if pw_err:
            return jsonify({"ok": False, "message": pw_err}), 400

    report_id = _parse_report_id(session.get("report_id"))
    if not report_id:
        return jsonify({"ok": False, "message": "Report not found. Please upload a dataset again."}), 400

    return _deliver_report_pdf_via_email(sender.email, report_id, sender=sender, pdf_password=pdf_password or None)


# ---------- PUBLIC FILES (for Twilio MediaUrl) ----------
@app.route('/files/<path:filename>', methods=['GET'])
def public_files(filename: str):
    """Serve generated report files for Twilio MediaUrl fetching.

    Twilio does not have access to user sessions/cookies, so this endpoint is intentionally
    unauthenticated. Files are keyed by report UUID and stored in a non-user-controlled
    directory.
    """
    name = (filename or "").strip()
    # Only allow our expected filenames.
    # Allow optional suffixes for derived protected files (e.g., public_<id>_pw.pdf, public_<id>_xlsx.zip).
    m = re.fullmatch(r"public_([0-9a-fA-F-]{32,36})(?:_([A-Za-z0-9]+))?\.(pdf|xlsx|zip)", name)
    if not m:
        return ("Not found", 404)

    report_id = m.group(1)
    suffix = m.group(2)
    ext = m.group(3)
    report_id = _parse_report_id(report_id)
    if not report_id:
        return ("Not found", 404)

    # For derived files (suffix present) and for all ZIPs, do not generate on demand.
    if suffix or ext == "zip":
        path = os.path.join(DOWNLOAD_DIR, name)
    else:
        # Base files public_<id>.pdf|xlsx can be generated on demand from cached report JSON.
        try:
            report_path = os.path.join(DOWNLOAD_DIR, f"{report_id}.json")
            with open(report_path, "r", encoding="utf-8") as f:
                report = json.load(f)
        except Exception:
            return ("Not found", 404)

        try:
            _ensure_public_report_files(report_id, report)
        except Exception:
            app.logger.exception("Failed generating public report files")
            return ("Not found", 404)

        path = os.path.join(DOWNLOAD_DIR, name)
    if not os.path.exists(path):
        return ("Not found", 404)

    # Do not cache long-term; these are ephemeral files.
    @after_this_request
    def _no_cache(resp):
        resp.headers["Cache-Control"] = "no-store"
        return resp

    if ext == "zip":
        return send_file(path, as_attachment=False, mimetype="application/zip")
    return send_file(path, as_attachment=False)


# ---------- SECURE EMAIL SHARING ----------

def _require_verified_sender() -> tuple[CurrentUser | None, tuple[dict, int] | None]:
    user = _sender_current_user_from_session()
    if user is None:
        return None, ({"ok": False, "message": "Please login first."}, 401)

    # Best-effort refresh if we don't have a verified flag yet.
    if not user.email_verified:
        _refresh_sender_verified_state_best_effort()
        user = _sender_current_user_from_session()

    if user is None or not user.email_verified:
        return None, ({"ok": False, "message": "Please verify your email first."}, 403)

    return user, None


def _validate_account_password_or_error(password: str) -> str | None:
    pw = (password or "").strip()
    if not pw:
        return "Password cannot be empty."
    if len(pw) < max(8, int(PW_CHANGE_MIN_PASSWORD_LEN)):
        return f"Password must be at least {max(8, int(PW_CHANGE_MIN_PASSWORD_LEN))} characters."
    if len(pw) > 128:
        return "Password is too long."
    if pw.lower() in COMMON_PASSWORDS:
        return "Password is too common. Choose a stronger password."
    return None


def _pw_change_otp_key(user_id: str) -> str:
    return "pwotp:" + hashlib.sha256((user_id or "").encode("utf-8")).hexdigest()


def _pw_change_otp_cooldown_key(user_id: str) -> str:
    return f"pwotp_cd:{(user_id or '').strip()}"


def _pw_change_store_otp_record(user_id: str, otp_hash: str) -> None:
    backend = _pw_store_backend()
    key = _pw_change_otp_key(user_id)
    record = {
        "otp_hash": otp_hash,
        "attempts": 0,
        "created_at": int(time.time()),
    }
    raw = json.dumps(record).encode("utf-8")
    if hasattr(backend, "setex"):
        backend.setex(key, int(PW_CHANGE_OTP_TTL_SECONDS), raw)
    else:
        backend.set(key, raw, int(PW_CHANGE_OTP_TTL_SECONDS))


def _pw_change_get_otp_record(user_id: str) -> dict | None:
    backend = _pw_store_backend()
    key = _pw_change_otp_key(user_id)
    try:
        raw = backend.get(key)
    except Exception:
        raw = None
    if not raw:
        return None
    try:
        obj = json.loads(raw.decode("utf-8"))
    except Exception:
        return None
    if not isinstance(obj, dict):
        return None
    return obj


def _pw_change_set_otp_record(user_id: str, record: dict) -> None:
    backend = _pw_store_backend()
    key = _pw_change_otp_key(user_id)
    raw = json.dumps(record).encode("utf-8")
    if hasattr(backend, "setex"):
        backend.setex(key, int(PW_CHANGE_OTP_TTL_SECONDS), raw)
    else:
        backend.set(key, raw, int(PW_CHANGE_OTP_TTL_SECONDS))


def _pw_change_delete_otp_record(user_id: str) -> None:
    backend = _pw_store_backend()
    key = _pw_change_otp_key(user_id)
    try:
        backend.delete(key)
    except Exception:
        return


def _otp_hash(value: str) -> str:
    v = (value or "").strip()
    # Add server secret so hashes aren't reusable across environments.
    material = f"{CONFIG.flask_secret_key}|{v}"
    return hashlib.sha256(material.encode("utf-8")).hexdigest()


def _supabase_auth_update_password(access_token: str, new_password: str) -> None:
    """Update password via Supabase Auth (requires a valid access token)."""
    if not CONFIG.supabase_url:
        raise RuntimeError("supabase_not_configured")
    url = (CONFIG.supabase_url or "").rstrip("/") + "/auth/v1/user"
    data = json.dumps({"password": new_password}).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        method="PUT",
        headers={
            "Authorization": f"Bearer {access_token}",
            "apikey": (CONFIG.supabase_anon_key or ""),
            "Content-Type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            status = getattr(resp, "status", None)
            if status not in (200, 204):
                raise RuntimeError("supabase_password_update_failed")
    except Exception as e:
        raise RuntimeError("supabase_password_update_failed") from e




def _send_email_plain_sendgrid(to_email: str, subject: str, body_text: str) -> None:
    cfg = _get_sendgrid_config()
    if not cfg.get("api_key") or not cfg.get("from_email"):
        raise RuntimeError("missing_sendgrid_config")

    payload = {
        "personalizations": [{"to": [{"email": to_email}]}],
        "from": {"email": cfg["from_email"]},
        "subject": subject,
        "content": [{"type": "text/plain", "value": body_text}],
    }

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.sendgrid.com/v3/mail/send",
        data=data,
        method="POST",
        headers={
            "Authorization": f"Bearer {cfg['api_key']}",
            "Content-Type": "application/json",
        },
    )

    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            status = getattr(resp, "status", None)
            if status not in (200, 202):
                raise RuntimeError("sendgrid_failed")
    except urllib.error.HTTPError as e:
        try:
            body_bytes = e.read()
            body_text = (body_bytes or b"").decode("utf-8", errors="replace")
        except Exception:
            body_text = ""
        app.logger.warning(
            "SendGrid HTTPError (plain) code=%s body=%s",
            getattr(e, "code", "unknown"),
            (body_text[:2000] if body_text else ""),
        )
        raise RuntimeError("sendgrid_failed") from e
    except urllib.error.URLError as e:
        app.logger.warning("SendGrid URLError (plain): %s", str(e))
        raise RuntimeError("sendgrid_failed") from e


def _app_public_base_url() -> str:
    # Use the configured email redirect base to generate public verification links.
    return (CONFIG.email_redirect_base_url or CONFIG.app_base_url or "").rstrip("/")


def _create_or_refresh_trusted_email_verification(owner_id: str, email: str, access_token: str) -> None:
    # Abuse prevention: limit verification email bursts per owner+email.
    rl_key = f"verify_trusted:{owner_id}:{email.lower().strip()}"
    if _rate_limit_hit(rl_key, VERIFY_EMAIL_RATE_LIMIT, VERIFY_EMAIL_RATE_WINDOW_SECONDS):
        raise RuntimeError("rate_limited")

    token = _new_verify_token()
    token_hash = _token_hash(token)

    sb = get_supabase(access_token)
    now_iso = datetime.now(timezone.utc).isoformat()
    row = _supabase_table_get_single(sb, "trusted_emails", {"owner_id": owner_id, "email": email})
    if row:
        sb.table("trusted_emails").update(
            {
                "verify_token_hash": token_hash,
                "verify_sent_at": now_iso,
                "is_verified": False,
                "verified_at": None,
            }
        ).eq("id", row["id"]).execute()
    else:
        sb.table("trusted_emails").insert(
            {
                "owner_id": owner_id,
                "email": email,
                "verify_token_hash": token_hash,
                "verify_sent_at": now_iso,
                "is_verified": False,
            }
        ).execute()

    verify_url = f"{_app_public_base_url()}/trusted-email/verify?token={token}"
    subject = "Verify your email for report delivery"
    body = (
        "You (or someone) added this email to receive reports.\n\n"
        f"Verify this email by opening: {verify_url}\n\n"
        "If you did not request this, you can ignore this message."
    )
    _send_email_plain_sendgrid(email, subject, body)


def _create_or_refresh_report_share_verification(
    owner_id: str,
    report_id: str,
    recipient_email: str,
    access_token: str,
    pdf_password: str | None = None,
) -> None:
    # Abuse prevention: limit verification email bursts and repeated spam.
    rl_key = f"verify_share:{owner_id}:{report_id}:{recipient_email.lower().strip()}"
    if _rate_limit_hit(rl_key, VERIFY_EMAIL_RATE_LIMIT, VERIFY_EMAIL_RATE_WINDOW_SECONDS):
        raise RuntimeError("rate_limited")
    cd_key = f"verify_share_cd:{owner_id}:{recipient_email.lower().strip()}"
    # Important: do NOT start cooldown until after the verification email is actually sent.
    if _cooldown_is_active(cd_key):
        raise RuntimeError("cooldown")

    recipient_email = (recipient_email or "").strip().lower()
    password_stored = True

    # If sender requested password protection, store password encrypted in short-lived storage.
    # Never persist the plaintext in the database.
    if pdf_password:
        try:
            _store_share_pdf_password(owner_id, report_id, recipient_email, pdf_password)
        except Exception:
            # Do not block verification email if password storage is unavailable.
            password_stored = False

    token = _new_verify_token()
    token_hash = _token_hash(token)
    now_dt = datetime.now(timezone.utc)
    now_iso = now_dt.isoformat()
    expires_dt = now_dt + timedelta(hours=24)

    sb = get_supabase(access_token)
    row = _supabase_table_get_single(
        sb,
        "report_shares",
        {"owner_id": owner_id, "report_id": report_id, "recipient_email": recipient_email},
    )

    payload = {
        "verify_token_hash": token_hash,
        "verify_sent_at": now_iso,
        "is_verified": False,
        "verified_at": None,
        "expires_at": expires_dt.isoformat(),
    }
    if row:
        sb.table("report_shares").update(payload).eq("id", row["id"]).execute()
    else:
        sb.table("report_shares").insert(
            {
                "owner_id": owner_id,
                "report_id": report_id,
                "recipient_email": recipient_email,
                **payload,
            }
        ).execute()

    verify_url = f"{_app_public_base_url()}/share/verify?token={token}"
    subject = "Verify your email to receive a shared report"
    body = (
        "A report was shared with this email address.\n\n"
        f"Verify ownership to allow delivery: {verify_url}\n\n"
        "If you did not expect this, you can ignore this message."
    )
    _send_email_plain_sendgrid(recipient_email, subject, body)

    # Start cooldown only after the email send succeeds.
    _cooldown_start(cd_key, RECIPIENT_COOLDOWN_SECONDS)

    if pdf_password and not password_stored:
        # Preserve previous behavior for callers that want to message the sender.
        raise RuntimeError("password_storage_unavailable")


def _load_report_from_cache(report_id: str) -> dict:
    rid = _parse_report_id(report_id)
    if not rid:
        raise FileNotFoundError("invalid_report_id")
    report_path = os.path.join(DOWNLOAD_DIR, f"{rid}.json")
    with open(report_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _deliver_cached_report_via_email(
    recipient_email: str,
    report_id: str,
    owner_id: str | None,
    pdf_password: str | None = None,
) -> None:
    """Deliver report email using cached report JSON (no session required)."""
    report = _load_report_from_cache(report_id)

    pdf_path = os.path.join(DOWNLOAD_DIR, f"retail_forecast_report_{report_id}.pdf")
    xlsx_path = os.path.join(DOWNLOAD_DIR, f"retail_forecast_report_{report_id}.xlsx")
    filename_pdf = f"retail_forecast_report_{report_id}.pdf"
    filename_xlsx = f"retail_forecast_report_{report_id}.xlsx"

    try:
        _build_pdf_report_file(report, pdf_path)
        _build_excel_report_file(report, xlsx_path)
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()
        with open(xlsx_path, "rb") as f:
            xlsx_bytes = f.read()

        pdf_password = (pdf_password or "").strip() if pdf_password is not None else ""
        subject = "Demand Forecasting Report"

        if pdf_password:
            pdf_bytes = _encrypt_pdf_bytes_if_requested(pdf_bytes, pdf_password)
            xlsx_open_pw_bytes = _encrypt_xlsx_bytes_password_to_open_or_none(xlsx_bytes, pdf_password)
            if xlsx_open_pw_bytes:
                body = "Attached are your password-protected report files (PDF and Excel)."
                delivered_kind = "pdf+xlsx_openpw"
                attachments = [
                    (filename_pdf, "application/pdf", pdf_bytes),
                    (filename_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", xlsx_open_pw_bytes),
                ]
            else:
                protected_xlsx_bytes = _protect_xlsx_bytes_if_requested(xlsx_bytes, pdf_password)
                body = "Attached are your password-protected report files (PDF and Excel)."
                delivered_kind = "pdf+xlsx_editpw"
                attachments = [
                    (filename_pdf, "application/pdf", pdf_bytes),
                    (filename_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", protected_xlsx_bytes),
                ]
        else:
            body = "Attached are your Retail Demand Forecasting report files (PDF and Excel)."
            delivered_kind = "pdf+xlsx"
            attachments = [
                (filename_pdf, "application/pdf", pdf_bytes),
                (filename_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", xlsx_bytes),
            ]

        _send_email_with_attachments_sendgrid(recipient_email, subject, body, attachments)
        _audit_event_best_effort(
            "report_delivered_auto",
            owner_id=owner_id,
            actor_email=None,
            recipient_email=recipient_email,
            report_id=report_id,
            meta={"password_protected": bool(pdf_password), "delivered": delivered_kind},
        )
    finally:
        try:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)
        except OSError:
            pass


def _audit_event_best_effort(
    event_type: str,
    owner_id: str | None,
    actor_email: str | None,
    recipient_email: str | None,
    report_id: str | None,
    meta: dict | None = None,
) -> None:
    """Best-effort audit logging.

    Security reasoning:
    - Audit logs help investigate abuse and delivery disputes.
    - Logging must not break user actions; failures are swallowed.
    - We use Supabase service role on the server (never exposed to clients).
    """
    payload = {
        "event_type": (event_type or "").strip()[:64],
        "owner_id": owner_id,
        "actor_email": (actor_email or "").strip().lower()[:254] if actor_email else None,
        "recipient_email": (recipient_email or "").strip().lower()[:254] if recipient_email else None,
        "report_id": report_id,
        "ip": _client_ip(),
        "user_agent": (request.headers.get("User-Agent") or "")[:512],
        "meta": meta or {},
    }
    try:
        sb_admin = get_supabase_admin()
        sb_admin.table("share_audit_events").insert(payload).execute()
    except Exception:
        try:
            app.logger.info("AUDIT %s %s", event_type, json.dumps(payload, default=str)[:2000])
        except Exception:
            pass


@app.route("/api/settings/password/init", methods=["POST"])
def api_settings_password_init():
    user, err = _require_verified_sender()
    if err:
        return jsonify(err[0]), err[1]

    access_token = (session.get("access_token") or "").strip()
    if not access_token:
        return jsonify({"ok": False, "message": "Login session expired. Please login again."}), 401

    payload = request.get_json(silent=True) or {}
    old_password = (payload.get("old_password") or payload.get("oldPassword") or "").strip()
    new_password = (payload.get("new_password") or payload.get("newPassword") or "").strip()
    confirm_password = (payload.get("confirm_password") or payload.get("confirmPassword") or "").strip()

    if not old_password:
        return jsonify({"ok": False, "message": "Old password is required."}), 400
    if not new_password or not confirm_password:
        return jsonify({"ok": False, "message": "New password and confirmation are required."}), 400
    if new_password != confirm_password:
        return jsonify({"ok": False, "message": "New password and confirmation do not match."}), 400

    pw_err = _validate_account_password_or_error(new_password)
    if pw_err:
        return jsonify({"ok": False, "message": pw_err}), 400

    cd_key = _pw_change_otp_cooldown_key(user.id)
    if _cooldown_is_active(cd_key):
        return jsonify({"ok": False, "message": "Please wait before requesting another OTP."}), 429

    # Verify old password by attempting login.
    if not UNIT_TESTING:
        try:
            sb = get_supabase()
            try:
                sb.auth.sign_in_with_password({"email": user.email, "password": old_password})
            except TypeError:
                sb.auth.sign_in_with_password(email=user.email, password=old_password)
        except Exception as e:
            msg = str(e).lower()
            if "missing supabase config" in msg or "supabase" in msg and "missing" in msg:
                return jsonify({"ok": False, "message": "Server is not configured for password changes."}), 500
            return jsonify({"ok": False, "message": "Old password is incorrect."}), 400

    otp = f"{secrets.randbelow(1000000):06d}"
    try:
        _pw_change_store_otp_record(user.id, _otp_hash(otp))
    except Exception:
        return jsonify({"ok": False, "message": "Failed to start OTP flow. Please try again."}), 500

    try:
        subject = "OTP to confirm password change"
        body = (
            "You requested a password change.\n\n"
            f"Your OTP code is: {otp}\n\n"
            f"This code expires in {max(1, int(PW_CHANGE_OTP_TTL_SECONDS // 60))} minutes.\n"
            "If you did not request this, you can ignore this email."
        )
        _send_email_plain_sendgrid(user.email, subject, body)
        _cooldown_start(cd_key, int(PW_CHANGE_OTP_COOLDOWN_SECONDS))
        return jsonify({"ok": True, "message": "OTP sent to your email."}), 200
    except Exception as e:
        _pw_change_delete_otp_record(user.id)
        msg = str(e).lower()
        if "missing_sendgrid_config" in msg:
            return jsonify({"ok": False, "message": "Email service is not configured on the server."}), 500
        return jsonify({"ok": False, "message": "Failed to send OTP. Please try again later."}), 500


@app.route("/api/settings/password/confirm", methods=["POST"])
def api_settings_password_confirm():
    user, err = _require_verified_sender()
    if err:
        return jsonify(err[0]), err[1]

    access_token = (session.get("access_token") or "").strip()
    if not access_token:
        return jsonify({"ok": False, "message": "Login session expired. Please login again."}), 401

    payload = request.get_json(silent=True) or {}
    otp = (payload.get("otp") or payload.get("code") or "").strip()
    new_password = (payload.get("new_password") or payload.get("newPassword") or "").strip()
    confirm_password = (payload.get("confirm_password") or payload.get("confirmPassword") or "").strip()

    if not otp:
        return jsonify({"ok": False, "message": "OTP is required."}), 400
    if not new_password or not confirm_password:
        return jsonify({"ok": False, "message": "New password and confirmation are required."}), 400
    if new_password != confirm_password:
        return jsonify({"ok": False, "message": "New password and confirmation do not match."}), 400

    pw_err = _validate_account_password_or_error(new_password)
    if pw_err:
        return jsonify({"ok": False, "message": pw_err}), 400

    record = _pw_change_get_otp_record(user.id)
    if not record:
        return jsonify({"ok": False, "message": "OTP is invalid or expired. Please request a new OTP."}), 400

    attempts = 0
    try:
        attempts = int(record.get("attempts") or 0)
    except Exception:
        attempts = 0

    if attempts >= int(PW_CHANGE_OTP_MAX_ATTEMPTS):
        _pw_change_delete_otp_record(user.id)
        return jsonify({"ok": False, "message": "Too many invalid attempts. Please request a new OTP."}), 400

    expected = (record.get("otp_hash") or "").strip()
    if not expected or not secrets.compare_digest(expected, _otp_hash(otp)):
        record["attempts"] = attempts + 1
        try:
            _pw_change_set_otp_record(user.id, record)
        except Exception:
            pass
        if (attempts + 1) >= int(PW_CHANGE_OTP_MAX_ATTEMPTS):
            _pw_change_delete_otp_record(user.id)
            return jsonify({"ok": False, "message": "Too many invalid attempts. Please request a new OTP."}), 400
        return jsonify({"ok": False, "message": "OTP is incorrect."}), 400

    try:
        _supabase_auth_update_password(access_token, new_password)
    except Exception:
        return jsonify({"ok": False, "message": "Failed to update password. Please try again later."}), 500

    _pw_change_delete_otp_record(user.id)
    session.clear()
    return jsonify({"ok": True, "message": "Password updated. Please login again.", "redirect_to": "/login"}), 200


@app.route("/api/trusted-emails", methods=["GET", "POST"])
def api_trusted_emails():
    user, err = _require_verified_sender()
    if err:
        return jsonify(err[0]), err[1]

    access_token = (session.get("access_token") or "").strip()
    if not access_token:
        return jsonify({"ok": False, "message": "Login session expired. Please login again."}), 401

    if request.method == "GET":
        if UNIT_TESTING:
            return jsonify({"ok": True, "items": []})
        sb = get_supabase(access_token)
        res = sb.table("trusted_emails").select("id,email,is_verified,created_at,verified_at").eq(
            "owner_id", user.id
        ).order("created_at", desc=True).execute()
        return jsonify({"ok": True, "items": getattr(res, "data", []) or []})

    payload = request.get_json(silent=True) or {}
    raw_email = payload.get("email")
    email, email_err = _normalize_and_validate_email_for_sending(raw_email)
    if email_err:
        return jsonify({"ok": False, "message": email_err}), 400

    if BLOCK_DISPOSABLE_EMAILS and _is_disposable_email_domain(email):
        _audit_event_best_effort(
            "trusted_email_blocked_disposable",
            owner_id=user.id,
            actor_email=user.email,
            recipient_email=email,
            report_id=None,
            meta={},
        )
        return jsonify({"ok": False, "message": "Disposable email domains are not allowed."}), 400

    try:
        _create_or_refresh_trusted_email_verification(user.id, email, access_token)
        _audit_event_best_effort(
            "trusted_verification_sent",
            owner_id=user.id,
            actor_email=user.email,
            recipient_email=email,
            report_id=None,
            meta={},
        )
        return jsonify({"ok": True, "message": "Verification email sent. Please verify to enable delivery."}), 202
    except Exception as e:
        msg = str(e).lower()
        if "missing_sendgrid_config" in msg:
            return jsonify({"ok": False, "message": "Email service is not configured on the server."}), 500
        if "rate_limited" in msg:
            return jsonify({"ok": False, "message": "Too many verification emails. Please try again later."}), 429
        app.logger.exception("Failed initiating trusted email verification")
        return jsonify({"ok": False, "message": "Failed to send verification email. Please try again later."}), 500


@app.route("/api/trusted-emails/<trusted_id>", methods=["DELETE"])
def api_trusted_emails_delete(trusted_id: str):
    user, err = _require_verified_sender()
    if err:
        return jsonify(err[0]), err[1]

    access_token = (session.get("access_token") or "").strip()
    if not access_token:
        return jsonify({"ok": False, "message": "Login session expired. Please login again."}), 401

    tid = (trusted_id or "").strip()
    if not tid:
        return jsonify({"ok": False, "message": "Invalid trusted email id."}), 400

    if UNIT_TESTING:
        return jsonify({"ok": True}), 200

    try:
        sb = get_supabase(access_token)
        res = sb.table("trusted_emails").delete().eq("id", tid).eq("owner_id", user.id).execute()
        deleted = getattr(res, "data", None)
        if isinstance(deleted, list) and len(deleted) == 0:
            return jsonify({"ok": False, "message": "Item not found."}), 404
        return jsonify({"ok": True}), 200
    except Exception:
        app.logger.exception("Failed deleting trusted email")
        return jsonify({"ok": False, "message": "Failed to delete trusted email."}), 500


@app.route("/api/report-shares", methods=["POST"])
def api_report_shares_create():
    user, err = _require_verified_sender()
    if err:
        return jsonify(err[0]), err[1]

    access_token = (session.get("access_token") or "").strip()
    if not access_token:
        return jsonify({"ok": False, "message": "Login session expired. Please login again."}), 401

    payload = request.get_json(silent=True) or {}
    raw_email = payload.get("recipient_email") or payload.get("email")
    pdf_password = (payload.get("pdf_password") or payload.get("password") or "")
    report_id = _parse_report_id(payload.get("report_id") or session.get("report_id"))
    if not report_id:
        return jsonify({"ok": False, "message": "Report not found. Please generate a report first."}), 400

    email, email_err = _normalize_and_validate_email_for_sending(raw_email)
    if email_err:
        return jsonify({"ok": False, "message": email_err}), 400

    if BLOCK_DISPOSABLE_EMAILS and _is_disposable_email_domain(email):
        _audit_event_best_effort(
            "share_blocked_disposable",
            owner_id=user.id,
            actor_email=user.email,
            recipient_email=email,
            report_id=report_id,
            meta={},
        )
        return jsonify({"ok": False, "message": "Disposable email domains are not allowed."}), 400

    pdf_password = (pdf_password or "").strip()
    if pdf_password:
        pw_err = _validate_pdf_password_or_error(pdf_password)
        if pw_err:
            return jsonify({"ok": False, "message": pw_err}), 400

    if email == user.email:
        return jsonify({"ok": False, "message": "Recipient is your own email. Use direct send instead."}), 400

    try:
        _create_or_refresh_report_share_verification(user.id, report_id, email, access_token, pdf_password=pdf_password or None)
        _audit_event_best_effort(
            "share_verification_sent",
            owner_id=user.id,
            actor_email=user.email,
            recipient_email=email,
            report_id=report_id,
            meta={"password_protected": bool(pdf_password)},
        )
        return jsonify({"ok": True, "message": "Verification email sent to recipient."}), 202
    except Exception as e:
        msg = str(e).lower()
        if "missing_sendgrid_config" in msg:
            return jsonify({"ok": False, "message": "Email service is not configured on the server."}), 500
        if "rate_limited" in msg:
            return jsonify({"ok": False, "message": "Too many verification emails. Please try again later."}), 429
        if "cooldown" in msg:
            return jsonify({"ok": False, "message": "Please wait before sending another verification email to this recipient."}), 429
        if "password_protection_not_configured" in msg or "password_storage_unavailable" in msg:
            _audit_event_best_effort(
                "share_verification_sent_password_not_stored",
                owner_id=user.id,
                actor_email=user.email,
                recipient_email=email,
                report_id=report_id,
                meta={"password_protected_requested": bool(pdf_password)},
            )
            return jsonify({
                "ok": True,
                "message": "Verification email sent. Password cannot be stored on this server; set the PDF password again when resending after verification.",
            }), 202
        app.logger.exception("Failed initiating report share verification")
        return jsonify({"ok": False, "message": "Failed to send verification email. Please try again later."}), 500


@app.route("/trusted-email/verify", methods=["GET"])
def verify_trusted_email():
    token = (request.args.get("token") or "").strip()
    if not token:
        _audit_event_best_effort(
            "trusted_verify_invalid",
            owner_id=None,
            actor_email=None,
            recipient_email=None,
            report_id=None,
            meta={"reason": "missing_token"},
        )
        return render_template("verify_email_share.html", message="Invalid verification link."), 400

    try:
        sb_admin = get_supabase_admin()
    except Exception:
        _audit_event_best_effort(
            "trusted_verify_failed",
            owner_id=None,
            actor_email=None,
            recipient_email=None,
            report_id=None,
            meta={"reason": "admin_client_unavailable"},
        )
        return render_template(
            "verify_email_share.html",
            message="Server is not configured for verification. Please contact support.",
        ), 500

    token_hash = _token_hash(token)
    try:
        res = (
            sb_admin.table("trusted_emails")
            .select("id")
            .eq("verify_token_hash", token_hash)
            .limit(1)
            .execute()
        )
        rows = getattr(res, "data", None) or []
        if not rows:
            _audit_event_best_effort(
                "trusted_verify_invalid",
                owner_id=None,
                actor_email=None,
                recipient_email=None,
                report_id=None,
                meta={"reason": "token_not_found"},
            )
            return render_template("verify_email_share.html", message="Verification link is invalid or expired."), 400

        row_id = rows[0]["id"]
        sb_admin.table("trusted_emails").update(
            {
                "is_verified": True,
                "verified_at": datetime.now(timezone.utc).isoformat(),
                "verify_token_hash": None,
            }
        ).eq("id", row_id).execute()

        _audit_event_best_effort(
            "trusted_verified",
            owner_id=None,
            actor_email=None,
            recipient_email=None,
            report_id=None,
            meta={"row_id": row_id},
        )
        return render_template("verify_email_share.html", message="Email verified successfully. You can now receive reports."), 200
    except Exception:
        app.logger.exception("Trusted email verification failed")
        _audit_event_best_effort(
            "trusted_verify_failed",
            owner_id=None,
            actor_email=None,
            recipient_email=None,
            report_id=None,
            meta={"reason": "exception"},
        )
        return render_template("verify_email_share.html", message="Verification failed. Please try again later."), 500


@app.route("/share/verify", methods=["GET"])
def verify_report_share():
    token = (request.args.get("token") or "").strip()
    if not token:
        _audit_event_best_effort(
            "share_verify_invalid",
            owner_id=None,
            actor_email=None,
            recipient_email=None,
            report_id=None,
            meta={"reason": "missing_token"},
        )
        return render_template("verify_email_share.html", message="Invalid verification link."), 400

    try:
        sb_admin = get_supabase_admin()
    except Exception:
        _audit_event_best_effort(
            "share_verify_failed",
            owner_id=None,
            actor_email=None,
            recipient_email=None,
            report_id=None,
            meta={"reason": "admin_client_unavailable"},
        )
        return render_template(
            "verify_email_share.html",
            message="Server is not configured for verification. Please contact support.",
        ), 500

    token_hash = _token_hash(token)
    try:
        res = (
            sb_admin.table("report_shares")
            .select("id,owner_id,report_id,recipient_email,expires_at")
            .eq("verify_token_hash", token_hash)
            .limit(1)
            .execute()
        )
        rows = getattr(res, "data", None) or []
        if not rows:
            _audit_event_best_effort(
                "share_verify_invalid",
                owner_id=None,
                actor_email=None,
                recipient_email=None,
                report_id=None,
                meta={"reason": "token_not_found"},
            )
            return render_template("verify_email_share.html", message="Verification link is invalid or expired."), 400

        row = rows[0]
        exp_dt = _parse_iso_datetime(row.get("expires_at")) if row.get("expires_at") else None
        if exp_dt and exp_dt < datetime.now(timezone.utc):
            _audit_event_best_effort(
                "share_verify_invalid",
                owner_id=None,
                actor_email=None,
                recipient_email=None,
                report_id=None,
                meta={"reason": "expired"},
            )
            return render_template("verify_email_share.html", message="Verification link is expired."), 400

        now_iso = datetime.now(timezone.utc).isoformat()
        # Important: clear expires_at on success. Token expiry is not share expiry.
        sb_admin.table("report_shares").update(
            {
                "is_verified": True,
                "verified_at": now_iso,
                "verify_token_hash": None,
                "expires_at": None,
            }
        ).eq("id", row["id"]).execute()

        # Promote the verified recipient into trusted_emails (best-effort) so future
        # reports can be delivered without requiring per-report verification.
        try:
            owner_id = row.get("owner_id")
            recipient_email = (row.get("recipient_email") or "").strip().lower()
            if owner_id and recipient_email:
                existing = (
                    sb_admin.table("trusted_emails")
                    .select("id")
                    .eq("owner_id", owner_id)
                    .eq("email", recipient_email)
                    .limit(1)
                    .execute()
                )
                existing_rows = getattr(existing, "data", None) or []
                if existing_rows:
                    sb_admin.table("trusted_emails").update(
                        {
                            "is_verified": True,
                            "verified_at": now_iso,
                            "verify_token_hash": None,
                        }
                    ).eq("id", existing_rows[0]["id"]).execute()
                else:
                    sb_admin.table("trusted_emails").insert(
                        {
                            "owner_id": owner_id,
                            "email": recipient_email,
                            "is_verified": True,
                            "verified_at": now_iso,
                        }
                    ).execute()
        except Exception:
            app.logger.info("Trusted email promotion skipped", exc_info=True)

        _audit_event_best_effort(
            "share_verified",
            owner_id=(str(row.get("owner_id")) if row.get("owner_id") else None),
            actor_email=None,
            recipient_email=(row.get("recipient_email") or None),
            report_id=(row.get("report_id") or None),
            meta={"row_id": row["id"]},
        )

        # Auto-deliver the report after verification (best-effort).
        try:
            owner_id = str(row.get("owner_id") or "").strip() or None
            report_id = _parse_report_id(row.get("report_id"))
            recipient_email = (row.get("recipient_email") or "").strip().lower()

            if owner_id and report_id and recipient_email:
                # If password protection was requested, only deliver automatically if we still
                # have the password in short-lived storage.
                password_required = _share_password_was_requested(owner_id, report_id, recipient_email)
                stored_pw = _get_share_pdf_password(owner_id, report_id, recipient_email)
                if password_required and not stored_pw:
                    return render_template(
                        "verify_email_share.html",
                        message="Recipient verified successfully. The report will be delivered by the sender (password is required and must be re-sent).",
                    ), 200

                _deliver_cached_report_via_email(recipient_email, report_id, owner_id, pdf_password=stored_pw or None)
                return render_template(
                    "verify_email_share.html",
                    message="Recipient verified successfully. The report has been sent to your email.",
                ), 200
        except FileNotFoundError:
            # Report cache may have been cleaned up.
            return render_template(
                "verify_email_share.html",
                message="Recipient verified successfully, but the report is no longer available. Please ask the sender to resend.",
            ), 200
        except Exception:
            app.logger.exception("Auto-delivery after share verification failed")
            return render_template(
                "verify_email_share.html",
                message="Recipient verified successfully, but automatic delivery failed. Please ask the sender to resend.",
            ), 200

        return render_template("verify_email_share.html", message="Recipient verified successfully."), 200
    except Exception:
        app.logger.exception("Report share verification failed")
        _audit_event_best_effort(
            "share_verify_failed",
            owner_id=None,
            actor_email=None,
            recipient_email=None,
            report_id=None,
            meta={"reason": "exception"},
        )
        return render_template("verify_email_share.html", message="Verification failed. Please try again later."), 500


@app.route("/trusted-emails/manage", methods=["GET"])
def trusted_emails_manage_page():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    # Require verified sender to view/manage the list.
    _user, err = _require_verified_sender()
    if err:
        # Preserve behavior consistency with API responses.
        return render_template("verify_email_share.html", message=err[0].get("message") or "Access denied."), err[1]

    return render_template("trusted_emails.html")


# ---------- SEND VIA WHATSAPP (Twilio Sandbox) ----------
@app.route('/send-whatsapp', methods=['POST'])
def send_whatsapp():
    if not session.get('logged_in'):
        return jsonify({"ok": False, "message": "Please login first."}), 401

    # Basic per-session rate limiting to avoid duplicate clicks.
    now = time.time()
    last = session.get("whatsapp_last_sent_at")
    try:
        if last is not None and (now - float(last)) < 15.0:
            return jsonify({"ok": False, "message": "Please wait a moment before sending again."}), 429
    except Exception:
        pass

    number = None
    pdf_password = None
    try:
        if request.is_json:
            payload = request.get_json(silent=True) or {}
            number = (payload.get('number') or payload.get('whatsapp') or '').strip()
            pdf_password = payload.get("pdf_password") or payload.get("password")
        else:
            number = (request.form.get('number') or '').strip()
            pdf_password = request.form.get("pdf_password") or request.form.get("password")
    except Exception:
        number = None
        pdf_password = None

    pdf_password = (pdf_password or "").strip() if pdf_password is not None else ""
    if pdf_password:
        pw_err = _validate_pdf_password_or_error(pdf_password)
        if pw_err:
            return jsonify({"ok": False, "message": pw_err}), 400

    if not _is_valid_whatsapp_number(number):
        return jsonify({"ok": False, "message": "Please enter a valid WhatsApp number (e.g., +91XXXXXXXXXX)."}), 400

    public_base = _public_https_base_url_or_none()
    if not public_base:
        return jsonify({
            "ok": False,
            "message": "Server is not configured with a public HTTPS base URL for WhatsApp media.",
        }), 500

    try:
        report = _load_report_from_session()
    except (RuntimeError, FileNotFoundError, json.JSONDecodeError):
        return jsonify({"ok": False, "message": "Report not found. Please upload a dataset again."}), 400

    report_id = session.get("report_id")
    report_id = _parse_report_id(report_id)
    if not report_id:
        return jsonify({"ok": False, "message": "Report not found. Please upload a dataset again."}), 400

    try:
        _ensure_public_report_files(report_id, report)

        if pdf_password:
            pdf_path = os.path.join(DOWNLOAD_DIR, f"public_{report_id}.pdf")
            xlsx_path = os.path.join(DOWNLOAD_DIR, f"public_{report_id}.xlsx")
            with open(pdf_path, "rb") as f:
                pdf_bytes = f.read()
            with open(xlsx_path, "rb") as f:
                xlsx_bytes = f.read()

            # Create a password-protected PDF as its own public file.
            protected_pdf_bytes = _encrypt_pdf_bytes_if_requested(pdf_bytes, pdf_password)
            protected_pdf_name = f"public_{report_id}_pw.pdf"
            protected_pdf_path = os.path.join(DOWNLOAD_DIR, protected_pdf_name)
            with open(protected_pdf_path, "wb") as f:
                f.write(protected_pdf_bytes)

            # Create a password-protected Excel as its own public file.
            # Prefer password-to-open on Windows+Excel when available, otherwise fall back
            # to edit-protected XLSX.
            xlsx_open_pw_bytes = _encrypt_xlsx_bytes_password_to_open_or_none(xlsx_bytes, pdf_password)
            protected_xlsx_bytes = xlsx_open_pw_bytes or _protect_xlsx_bytes_if_requested(xlsx_bytes, pdf_password)
            protected_xlsx_name = f"public_{report_id}_pw.xlsx"
            protected_xlsx_path = os.path.join(DOWNLOAD_DIR, protected_xlsx_name)
            with open(protected_xlsx_path, "wb") as f:
                f.write(protected_xlsx_bytes)

            protected_pdf_url = f"{public_base}/files/{protected_pdf_name}"
            protected_xlsx_url = f"{public_base}/files/{protected_xlsx_name}"

            # Send as two separate WhatsApp messages for compatibility.
            sid_pdf = _send_whatsapp_media_twilio(number, protected_pdf_url, "Your password-protected Report (PDF)")
            app.logger.info("Twilio WhatsApp protected PDF sent sid=%s to=%s", sid_pdf, number)
            sid_xlsx = _send_whatsapp_media_twilio(number, protected_xlsx_url, "Your password-protected Report (Excel)")
            app.logger.info("Twilio WhatsApp protected Excel sent sid=%s to=%s", sid_xlsx, number)
            sids = [sid_pdf, sid_xlsx]
        else:
            pdf_url = f"{public_base}/files/public_{report_id}.pdf"
            xlsx_url = f"{public_base}/files/public_{report_id}.xlsx"

            # Send as two separate WhatsApp messages to maximize compatibility.
            sid_pdf = _send_whatsapp_media_twilio(number, pdf_url, "Your Demand Forecasting Report (PDF)")
            app.logger.info("Twilio WhatsApp PDF sent sid=%s to=%s", sid_pdf, number)
            sid_xlsx = _send_whatsapp_media_twilio(number, xlsx_url, "Your Demand Forecasting Report (Excel)")
            app.logger.info("Twilio WhatsApp Excel sent sid=%s to=%s", sid_xlsx, number)
            sids = [sid_pdf, sid_xlsx]

        session["whatsapp_last_sent_at"] = now
        return jsonify({"ok": True, "message": "Report sent to WhatsApp successfully.", "sids": sids})
    except Exception as e:
        # Log diagnostic details without leaking secrets.
        twilio_code = getattr(e, "code", None)
        twilio_status = getattr(e, "status", None)
        twilio_msg = getattr(e, "msg", None)
        safe_text = str(twilio_msg or e or "")
        app.logger.exception(
            "Twilio WhatsApp send failed type=%s code=%s status=%s msg=%s",
            type(e).__name__,
            twilio_code,
            twilio_status,
            (safe_text[:500] if safe_text else ""),
        )
        msg = str(e).lower()
        if "missing_twilio_config" in msg:
            return jsonify({"ok": False, "message": "WhatsApp service is not configured on the server."}), 500
        if "missing_twilio_sdk" in msg:
            return jsonify({"ok": False, "message": "WhatsApp service is unavailable on the server."}), 500
        if "zip_encryption_unavailable" in msg or "xlsx_protection_unavailable" in msg:
            return jsonify({"ok": False, "message": "Excel password protection is not available on the server."}), 503
        return jsonify({"ok": False, "message": _twilio_error_to_user_message(e)}), 500

if __name__ == "__main__":
    app.run(debug=True)
